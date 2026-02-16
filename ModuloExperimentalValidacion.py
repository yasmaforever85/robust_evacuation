#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ModuloExperimentalValidacion.py
═══════════════════════════════════════════════════════════════════════
Módulo Experimental de Validación para Generación de Escenarios
de Huracanes — Figuras de Calidad Publicación + Excel Enriquecido.

Este módulo extiende la plataforma de generación de escenarios
(GeneradorEscenariosHuracan.py) con:

  1. Figuras PDF de calidad publicación para el artículo JUS
     (no disponibles en la clase Charts del GUI)
  2. Libro Excel enriquecido con hojas de análisis adicionales:
     bootstrap, entropía cruzada, comparación multi-huracán
  3. Ejecución independiente (NO requiere el módulo estocástico PI^S)

Figuras generadas:
  fig0_framework.pdf        — Diagrama del pipeline computacional
  fig2_silhouette_bic.pdf   — Bootstrap: silhouette box plots + BIC
  fig4_prob_convergence.pdf — Convergencia multi-huracán (2 paneles)
  fig6_impact_map.pdf       — Cascada geográfica Ian 5 horizontes
  fig_uncertainty_ab.pdf    — Cascada bayesiana + conjunto incertidumbre
  fig_entropy_multi.pdf     — Entropía por huracán y horizonte
  fig_cat_heatmap.pdf       — Heatmap categoría × horizonte × huracán
  fig_demand_evolution.pdf  — Factor demanda ponderado multi-huracán

Uso:
  python3 ModuloExperimentalValidacion.py

Salida:
  pub_figures/               — Directorio con 8 figuras PDF
  escenarios_validacion.xlsx — Excel enriquecido (10 hojas)

Dependencias: GeneradorEscenariosHuracan.py, matplotlib, numpy, openpyxl
Autor: Yasmany Fernández-Fernández  |  Fecha: 2025
═══════════════════════════════════════════════════════════════════════
"""

import os
import sys
import math
import time
import numpy as np
from datetime import datetime
from typing import Dict, List, Optional, Tuple

# ─── Matplotlib config para publicación ─────────────────────────────
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
from matplotlib.lines import Line2D

# Estilo publicación global
plt.rcParams.update({
    'font.family': 'serif',
    'font.size': 10,
    'axes.labelsize': 11,
    'axes.titlesize': 12,
    'legend.fontsize': 9,
    'xtick.labelsize': 9,
    'ytick.labelsize': 9,
    'figure.dpi': 150,
    'savefig.dpi': 300,
    'savefig.bbox': 'tight',
    'axes.grid': True,
    'grid.alpha': 0.3,
    'axes.spines.top': False,
    'axes.spines.right': False,
})

# ─── Importar motor de escenarios ───────────────────────────────────
from GeneradorEscenariosHuracan import ScenarioGenerator, ScenarioSet

# ═════════════════════════════════════════════════════════════════════
# CONSTANTES
# ═════════════════════════════════════════════════════════════════════

HURRICANES = ['IAN', 'IRMA', 'SANDY', 'MATTHEW']
HORIZONS = [72, 48, 36, 24, 12]
SEED = 42
N_BOOTSTRAP = 20

# Paletas
HCOLORS = {
    'IAN': '#E74C3C', 'IRMA': '#3498DB',
    'SANDY': '#2ECC71', 'MATTHEW': '#9B59B6'
}
HMARKERS = {'IAN': 's', 'IRMA': 'o', 'SANDY': '^', 'MATTHEW': 'D'}
HORIZON_COLORS = {72: '#E74C3C', 48: '#F39C12', 36: '#27AE60',
                  24: '#2980B9', 12: '#8B008B'}
CAT_COLORS = {1: '#3498DB', 2: '#2ECC71', 3: '#F1C40F',
              4: '#E67E22', 5: '#E74C3C'}

# Costas de Cuba (simplificada para mapa)
CUBA_COAST_N = [
    (-84.95, 21.86), (-84.50, 22.03), (-83.90, 22.40), (-83.50, 22.70),
    (-83.10, 22.91), (-82.70, 23.04), (-82.30, 23.10), (-81.80, 23.15),
    (-81.20, 23.05), (-80.60, 22.90), (-80.00, 22.70), (-79.50, 22.50),
    (-79.20, 22.30), (-78.80, 22.10), (-78.00, 21.85), (-77.50, 21.60),
    (-77.10, 21.50), (-76.50, 20.90), (-76.20, 20.40), (-75.80, 20.05),
    (-74.80, 20.10)
]
CUBA_COAST_S = [
    (-84.95, 21.86), (-84.20, 21.58), (-83.70, 21.65), (-83.00, 21.80),
    (-82.40, 21.70), (-82.10, 21.50), (-81.60, 21.65), (-81.00, 21.55),
    (-80.50, 21.50), (-80.00, 21.45), (-79.50, 21.60), (-79.00, 21.50),
    (-78.50, 21.30), (-78.00, 20.90), (-77.50, 20.50), (-77.00, 20.30),
    (-76.50, 19.90), (-75.80, 20.05), (-74.80, 20.10)
]


# ═════════════════════════════════════════════════════════════════════
# CLASE PRINCIPAL: GeneradorFigurasPublicacion
# ═════════════════════════════════════════════════════════════════════

class GeneradorFigurasPublicacion:
    """
    Genera figuras de calidad publicación a partir de datos
    del GeneradorEscenariosHuracan — sin necesidad del solver PI^S.
    """

    def __init__(self, output_dir: str = "pub_figures"):
        self.gen = ScenarioGenerator()
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

        # Cache de cascadas {hurricane: {horizon: ScenarioSet}}
        self.cascades: Dict[str, Dict[int, ScenarioSet]] = {}
        # Cache de bootstrap {hurricane: {horizon: {seed: ScenarioSet}}}
        self.bootstrap: Dict[str, Dict[int, Dict[int, ScenarioSet]]] = {}

    # ─── Generación de datos ────────────────────────────────────────

    def generate_all_data(self):
        """Genera cascadas para todos los huracanes y horizontes."""
        print("═" * 60)
        print("  GENERANDO DATOS DE ESCENARIOS")
        print("═" * 60)

        for hurr in HURRICANES:
            self.cascades[hurr] = {}
            for h in HORIZONS:
                ss = self.gen.generate_scenarios(
                    hurr, h, method="gmm", seed=SEED)
                self.cascades[hurr][h] = ss
            n_sc = len(self.cascades[hurr][72].scenarios)
            p_dom = max(s.probability
                        for s in self.cascades[hurr][72].scenarios)
            print(f"  {hurr:8s}: S={n_sc}, p*(72h)={p_dom:.3f}")

        # Bootstrap
        print("\n  Bootstrap (B=%d)..." % N_BOOTSTRAP)
        for hurr in HURRICANES:
            self.bootstrap[hurr] = {}
            for h in HORIZONS:
                self.bootstrap[hurr][h] = {}
                for b in range(N_BOOTSTRAP):
                    ss = self.gen.generate_scenarios(
                        hurr, h, method="gmm", seed=b)
                    self.bootstrap[hurr][h][b] = ss
        print("  ✓ Bootstrap completo\n")

    # ─── Utilidades ─────────────────────────────────────────────────

    def _entropy(self, ss: ScenarioSet) -> float:
        return -sum(s.probability * math.log2(s.probability)
                    for s in ss.scenarios if s.probability > 1e-15)

    def _save(self, fig, name: str):
        path = os.path.join(self.output_dir, name)
        fig.savefig(path, bbox_inches='tight', dpi=300)
        plt.close(fig)
        print(f"  ✓ {name}")

    def _draw_cuba(self, ax):
        """Dibuja contorno simplificado de Cuba."""
        lons_n = [p[0] for p in CUBA_COAST_N]
        lats_n = [p[1] for p in CUBA_COAST_N]
        lons_s = [p[0] for p in CUBA_COAST_S]
        lats_s = [p[1] for p in CUBA_COAST_S]
        ax.fill(lons_n + lons_s[::-1],
                lats_n + lats_s[::-1],
                alpha=0.12, color='#2C3E50', zorder=1)
        ax.plot(lons_n, lats_n, '-', color='#7F8C8D',
                linewidth=1.2, alpha=0.6, zorder=2)
        ax.plot(lons_s, lats_s, '-', color='#7F8C8D',
                linewidth=1.2, alpha=0.6, zorder=2)

    # ═════════════════════════════════════════════════════════════════
    # FIGURA 1: Framework Computacional
    # ═════════════════════════════════════════════════════════════════

    def fig_framework(self):
        """Pipeline diagram del sistema de generación de escenarios."""
        fig, ax = plt.subplots(figsize=(14, 8.5))
        ax.set_xlim(0, 14); ax.set_ylim(0, 8.5)
        ax.axis('off')
        ax.set_facecolor('white'); fig.patch.set_facecolor('white')

        ax.text(7, 8.2, 'COMPUTATIONAL FRAMEWORK',
                fontsize=16, fontweight='bold', ha='center')
        ax.text(7, 7.85,
                'Hurricane Scenario Generation & '
                'Impact Probability Estimation',
                fontsize=11, ha='center', style='italic')

        def box(x, y, w, h, title, text,
                color='#E8F0FE', ec='#4285F4', lw=1.5):
            rect = mpatches.FancyBboxPatch(
                (x - w/2, y - h/2), w, h,
                boxstyle="round,pad=0.1",
                fc=color, ec=ec, lw=lw)
            ax.add_patch(rect)
            ax.text(x, y + h/4, title, fontsize=9,
                    fontweight='bold', ha='center', va='center')
            ax.text(x, y - h/6, text, fontsize=7,
                    ha='center', va='center', color='#333')

        def arrow(x1, y1, x2, y2):
            ax.annotate('', xy=(x2, y2), xytext=(x1, y1),
                        arrowprops=dict(arrowstyle='->',
                                        color='#555', lw=1.5))

        # Stage labels
        for sy, lbl in [(6.7, 'INPUT'), (5.15, 'STAGE 1'),
                        (3.65, 'STAGE 2'), (1.8, 'OUTPUT')]:
            ax.text(0.35, sy, lbl, fontsize=8, fontweight='bold',
                    rotation=90, ha='center', va='center', color='#666',
                    bbox=dict(facecolor='#F0F0F0', edgecolor='#CCC',
                              boxstyle='round'))

        # INPUT
        box(2.5, 7.0, 2.8, 0.9, 'HURDAT2 Database',
            'NHC Best Track\n21 storms (1926–2024)')
        box(5.5, 7.0, 2.8, 0.9, 'Target Hurricane',
            'Name, category,\ncurrent position')
        box(8.5, 7.0, 2.8, 0.9, 'Forecast Horizon',
            r'$h \in \{72,48,36,24,12\}$' + '\nhours before landfall')
        box(11.5, 7.0, 2.4, 0.9, 'Manual Storms',
            'CSV import\n(current season)')

        # STAGE 1
        box(4, 5.4, 4.5, 0.9,
            'Cuba Filter + 9-D Vectorization',
            r'$\varphi(i)$ = [cat, wind/185, dist/500, land,'
            '\ntransit/48, bearing/360, reg, 2×lat, 2×lon]',
            color='#FFF3E0', ec='#FF9800')
        box(10, 5.4, 3.8, 0.9,
            'Analog Search (Euclidean in 9-D)',
            r'Top-N nearest neighbors'
            '\n' + r'Geographic amplification $\kappa_g = 2.0$',
            color='#FFF3E0', ec='#FF9800')

        # STAGE 2
        box(2.5, 3.8, 3.2, 0.9, 'GMM + BIC Selection',
            r'$k^* = \arg\min$ BIC($k$), $k = 2..7$'
            '\nEM algorithm, K-Means++ init',
            color='#E8F5E9', ec='#4CAF50')
        box(7, 3.8, 3.2, 0.9, 'Bayesian Updating',
            r'$P(\omega_s|h) \propto \pi_s \cdot '
            r'\exp(-\|\mu-q\|^2 / 2\sigma^2(h))$'
            '\n' + r'$\sigma^2$: 1.0 → 0.6 → 0.4 → 0.25 → 0.15',
            color='#E8F5E9', ec='#4CAF50')
        box(11, 3.8, 2.8, 0.9, 'Impact Geolocation',
            'Cluster centroid (lat, lon)\n'
            'Haversine → nearest locality',
            color='#E8F5E9', ec='#4CAF50')

        # Arrows
        for x in [2.5, 5.5, 8.5]:
            arrow(x, 6.5, x if x < 6 else x - 1, 5.9)
        arrow(11.5, 6.5, 10, 5.9)
        arrow(5.5, 4.9, 4, 4.3)
        arrow(8, 4.9, 7, 4.3)
        arrow(10, 4.9, 11, 4.3)
        arrow(2.5, 3.3, 5, 2.5)
        arrow(7, 3.3, 7, 2.5)
        arrow(11, 3.3, 9, 2.5)

        # Compute real entropy values
        ent_72 = np.mean([self._entropy(self.cascades[h][72])
                          for h in HURRICANES])
        ent_12 = np.mean([self._entropy(self.cascades[h][12])
                          for h in HURRICANES])
        red = (1 - ent_12 / ent_72) * 100

        # OUTPUT
        out = mpatches.FancyBboxPatch(
            (1.2, 1.3), 11.6, 1.8,
            boxstyle="round,pad=0.15",
            fc='#FDE2E2', ec='#C0392B', lw=2)
        ax.add_patch(out)
        ax.text(7, 2.65,
                r'OUTPUT:  Scenario Set  '
                r'$\Xi(\tau) = \{\omega_1, \omega_2, ..., \omega_k\}$',
                fontsize=12, fontweight='bold', ha='center')
        ax.text(7, 2.1,
                r'Each $\omega_s$ = ( $P(\omega_s|\tau)$, Category, '
                'Wind, Demand, Lat, Lon, Locality, Province )',
                fontsize=9, ha='center')
        ax.text(7, 1.6,
                r'$\sum_s P(\omega_s|\tau) = 1.000$   |   '
                r'$\bar{H}(72h) \approx %.1f$ bits  →  '
                r'$\bar{H}(12h) \approx %.1f$ bits  '
                r'(%d%% reduction)' % (ent_72, ent_12, red),
                fontsize=9, ha='center', color='#C0392B',
                fontweight='bold')

        self._save(fig, 'fig0_framework.pdf')

    # ═════════════════════════════════════════════════════════════════
    # FIGURA 2: Silhouette Bootstrap + BIC
    # ═════════════════════════════════════════════════════════════════

    def fig_silhouette_bic(self):
        """Box plots bootstrap de silhouette + BIC por horizonte."""
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4.5))

        # (a) Silhouette box plots
        sil_data = []
        labels = []
        for h in HORIZONS:
            sils = []
            for hurr in HURRICANES:
                for b in range(N_BOOTSTRAP):
                    ss = self.bootstrap[hurr][h][b]
                    if ss.silhouette_score is not None:
                        sils.append(ss.silhouette_score)
            sil_data.append(sils)
            labels.append('%dh' % h)

        bp = ax1.boxplot(sil_data, tick_labels=labels, patch_artist=True,
                         widths=0.5,
                         medianprops=dict(color='black', linewidth=2))
        for patch, h in zip(bp['boxes'], HORIZONS):
            patch.set_facecolor(HORIZON_COLORS[h])
            patch.set_alpha(0.7)
        ax1.axhline(y=0.25, color='#E74C3C', linestyle='--',
                    linewidth=1.5, alpha=0.7, label=r'$\bar{s} = 0.25$')
        ax1.set_xlabel('Forecast horizon')
        ax1.set_ylabel('Silhouette score')
        ax1.set_title('(a) Cluster quality (bootstrap B=%d)' % N_BOOTSTRAP)
        ax1.legend(fontsize=9)

        # (b) BIC by k for each horizon (Ian, seed=42)
        for h in HORIZONS:
            bics = {}
            for k in range(2, 8):
                ss = self.gen.generate_scenarios(
                    'IAN', h, method='gmm', seed=SEED, n_scenarios=k)
                if ss.bic is not None:
                    bics[k] = ss.bic
            if bics:
                ks = sorted(bics.keys())
                ax2.plot(ks, [bics[k] for k in ks],
                         '-o', color=HORIZON_COLORS[h],
                         linewidth=1.8, markersize=6,
                         label='%dh' % h)
        ax2.set_xlabel('Number of components $k$')
        ax2.set_ylabel('BIC')
        ax2.set_title('(b) BIC model selection (Ian)')
        ax2.legend(fontsize=8, ncol=2)

        fig.tight_layout()
        self._save(fig, 'fig2_silhouette_bic.pdf')

    # ═════════════════════════════════════════════════════════════════
    # FIGURA 4: Convergencia Probabilidad Multi-Huracán
    # ═════════════════════════════════════════════════════════════════

    def fig_prob_convergence(self):
        """2 paneles: (a) p*(τ) dominante, (b) escenarios activos."""
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4.5))

        for hurr in HURRICANES:
            p_doms, n_active = [], []
            for h in HORIZONS:
                ss = self.cascades[hurr][h]
                probs = [s.probability for s in ss.scenarios]
                p_doms.append(max(probs))
                n_active.append(sum(1 for p in probs if p > 0.01))

            ax1.plot(range(len(HORIZONS)), p_doms,
                     '-' + HMARKERS[hurr], color=HCOLORS[hurr],
                     linewidth=2, markersize=8, label=hurr)
            ax2.plot(range(len(HORIZONS)), n_active,
                     '-' + HMARKERS[hurr], color=HCOLORS[hurr],
                     linewidth=2, markersize=8, label=hurr)

        ax1.set_xticks(range(len(HORIZONS)))
        ax1.set_xticklabels(['%dh' % h for h in HORIZONS])
        ax1.set_xlabel('Forecast horizon $\\tau$')
        ax1.set_ylabel('$P(\\omega^*|\\tau)$')
        ax1.set_title('(a) Dominant scenario probability')
        ax1.set_ylim(0.3, 1.05)
        ax1.legend(fontsize=9)

        ax2.set_xticks(range(len(HORIZONS)))
        ax2.set_xticklabels(['%dh' % h for h in HORIZONS])
        ax2.set_xlabel('Forecast horizon $\\tau$')
        ax2.set_ylabel('Active scenarios ($p_s > 0.01$)')
        ax2.set_title('(b) Number of active scenarios')
        ax2.set_ylim(0, 6)
        ax2.legend(fontsize=9)

        fig.tight_layout()
        self._save(fig, 'fig4_prob_convergence.pdf')

    # ═════════════════════════════════════════════════════════════════
    # FIGURA 6: Mapa de Impacto Cascada Geográfica
    # ═════════════════════════════════════════════════════════════════

    def fig_impact_map(self):
        """Mapa Cuba con cascada geográfica de Ian a 5 horizontes."""
        fig, ax = plt.subplots(figsize=(11, 5.5))
        self._draw_cuba(ax)

        # Landfall real de Ian (Pinar del Río)
        ax.plot(-83.7, 22.4, '*', color='red', markersize=15,
                markeredgecolor='black', markeredgewidth=0.8,
                zorder=10, label='Actual landfall')

        for h in HORIZONS:
            ss = self.cascades['IAN'][h]
            for s in ss.scenarios:
                if s.impact_lat == 0 and s.impact_lon == 0:
                    continue
                size = max(30, s.probability * 1500)
                color = HORIZON_COLORS[h]
                alpha = 0.4 + 0.4 * s.probability
                ax.scatter(s.impact_lon, s.impact_lat, s=size,
                           c=color, edgecolors='black', linewidth=0.5,
                           alpha=alpha, zorder=5)
                if s.probability > 0.15:
                    ax.annotate(
                        '$\\omega_{%d}$\n%s\n%.0f%%' % (
                            s.id, s.impact_location,
                            s.probability * 100),
                        (s.impact_lon, s.impact_lat),
                        textcoords="offset points", xytext=(12, 8),
                        fontsize=7, fontweight='bold',
                        bbox=dict(boxstyle="round,pad=0.2",
                                  fc="white", alpha=0.85),
                        arrowprops=dict(arrowstyle="-",
                                        color="gray", alpha=0.5))

        # Legend for horizons
        handles = [Line2D([0], [0], marker='o', color='w',
                          markerfacecolor=HORIZON_COLORS[h],
                          markeredgecolor='black', markersize=8,
                          label='%dh' % h) for h in HORIZONS]
        handles.append(Line2D([0], [0], marker='*', color='w',
                              markerfacecolor='red',
                              markeredgecolor='black', markersize=12,
                              label='Landfall'))
        ax.legend(handles=handles, fontsize=8, loc='lower left',
                  framealpha=0.9)

        ax.set_xlabel('Longitude')
        ax.set_ylabel('Latitude')
        ax.set_title('Hurricane Ian — Scenario impact points '
                     'across forecast horizons')
        ax.set_xlim(-86, -74); ax.set_ylim(19.5, 24.0)

        fig.tight_layout()
        self._save(fig, 'fig6_impact_map.pdf')

    # ═════════════════════════════════════════════════════════════════
    # FIGURA 5 (paneles a,b): Incertidumbre — cascada + conjunto Ξ
    # ═════════════════════════════════════════════════════════════════

    def fig_uncertainty_ab(self):
        """2 paneles: (a) cascada bayesiana Ian, (b) conjunto Ξ(72)."""
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5))

        # (a) Bayesian cascade — Ian, 5 scenarios across horizons
        ss72 = self.cascades['IAN'][72]
        scenario_ids = [s.id for s in ss72.scenarios]
        colors_sc = ['#E74C3C', '#2980B9', '#27AE60',
                     '#F39C12', '#8B008B']

        for idx, sid in enumerate(scenario_ids):
            probs = []
            for h in HORIZONS:
                p = 0
                for s in self.cascades['IAN'][h].scenarios:
                    if s.id == sid:
                        p = s.probability
                        break
                probs.append(p)

            s72 = [s for s in ss72.scenarios if s.id == sid][0]
            label = ('$\\omega_{%d}$ %s Cat%d' %
                     (sid, s72.impact_location[:10], s72.category))
            ax1.plot(range(len(HORIZONS)), probs,
                     '-o', color=colors_sc[idx % 5],
                     linewidth=2, markersize=7, label=label)

        ax1.set_xticks(range(len(HORIZONS)))
        ax1.set_xticklabels(['%dh' % h for h in HORIZONS])
        ax1.set_xlabel('Forecast horizon $\\tau$ (h)')
        ax1.set_ylabel('$P(\\omega_s|\\tau)$')
        ax1.set_title('(a) Bayesian probability cascade — '
                      'Hurricane Ian')
        ax1.set_ylim(-0.02, 1.05)
        ax1.legend(fontsize=7, loc='upper right')

        # (b) Uncertainty set Ξ(72) — all hurricanes
        # P_nom = 60, Σπ = 75
        P_NOM = 60
        SUM_PI = 75
        ax2.axhline(y=SUM_PI, color='#C0392B', linestyle='--',
                    linewidth=2, alpha=0.8,
                    label='$\\Sigma\\pi = %d$' % SUM_PI)
        ax2.axhspan(SUM_PI, SUM_PI + 20, alpha=0.08,
                    color='#E74C3C', label='Infeasibility region')

        for hurr in HURRICANES:
            ss = self.cascades[hurr][72]
            for s in ss.scenarios:
                Ps = int(round(s.demand_factor * P_NOM))
                size = max(40, s.probability * 800)
                ax2.scatter(s.probability, Ps, s=size,
                            c=HCOLORS[hurr],
                            edgecolors='black', linewidth=0.5,
                            alpha=0.8, zorder=5)

            # Mean demand (P_ev)
            P_ev = sum(s.probability *
                       int(round(s.demand_factor * P_NOM))
                       for s in ss.scenarios)
            ax2.plot([], [], 's', color=HCOLORS[hurr],
                     markersize=8, label='%s ($\\bar{P}_{EV}$=%.0f)'
                     % (hurr, P_ev))

        ax2.set_xlabel('Scenario probability $p_s$')
        ax2.set_ylabel('Scenario demand $P_s$ (persons)')
        ax2.set_title('(b) Uncertainty set $\\Xi^{(72)}$ vs '
                      'shelter capacity')
        ax2.legend(fontsize=7, loc='upper right')

        fig.tight_layout()
        self._save(fig, 'fig_uncertainty_ab.pdf')

    # ═════════════════════════════════════════════════════════════════
    # FIGURA: Entropía Multi-Huracán
    # ═════════════════════════════════════════════════════════════════

    def fig_entropy_multi(self):
        """Entropía por horizonte para cada huracán + promedio."""
        fig, ax = plt.subplots(figsize=(8, 5))

        x = np.arange(len(HORIZONS))
        width = 0.18

        for i, hurr in enumerate(HURRICANES):
            ents = [self._entropy(self.cascades[hurr][h])
                    for h in HORIZONS]
            bars = ax.bar(x + i * width, ents, width,
                          color=HCOLORS[hurr], edgecolor='white',
                          label=hurr, alpha=0.85)
            for bar, e in zip(bars, ents):
                if e > 0.05:
                    ax.text(bar.get_x() + bar.get_width()/2,
                            bar.get_height() + 0.03,
                            '%.2f' % e, ha='center', fontsize=7,
                            fontweight='bold')

        # Average line
        mean_ents = [np.mean([self._entropy(self.cascades[h2][h])
                              for h2 in HURRICANES])
                     for h in HORIZONS]
        ax.plot(x + 1.5 * width, mean_ents, 'k--D', linewidth=2,
                markersize=6, label='Mean', zorder=10)

        ax.set_xticks(x + 1.5 * width)
        ax.set_xticklabels(['%dh' % h for h in HORIZONS])
        ax.set_xlabel('Forecast horizon $\\tau$')
        ax.set_ylabel('Shannon entropy $H(\\Xi)$ (bits)')
        ax.set_title('Entropy reduction across forecast horizons')
        ax.legend(fontsize=9)
        ax.set_ylim(0, max(mean_ents) * 1.3)

        fig.tight_layout()
        self._save(fig, 'fig_entropy_multi.pdf')

    # ═════════════════════════════════════════════════════════════════
    # FIGURA: Heatmap Categoría × Horizonte
    # ═════════════════════════════════════════════════════════════════

    def fig_cat_heatmap(self):
        """Heatmap de probabilidad por categoría para cada huracán."""
        fig, axes = plt.subplots(2, 2, figsize=(12, 8))
        axes = axes.flatten()

        for idx, hurr in enumerate(HURRICANES):
            ax = axes[idx]
            cascade = self.cascades[hurr]
            all_cats = sorted(set(
                s.category for ss in cascade.values()
                for s in ss.scenarios))

            matrix = np.zeros((len(all_cats), len(HORIZONS)))
            for j, h in enumerate(HORIZONS):
                for s in cascade[h].scenarios:
                    i = all_cats.index(s.category)
                    matrix[i, j] += s.probability

            im = ax.imshow(matrix, aspect='auto', cmap='YlOrRd',
                           interpolation='nearest',
                           vmin=0, vmax=1)
            ax.set_xticks(range(len(HORIZONS)))
            ax.set_xticklabels(['%dh' % h for h in HORIZONS])
            ax.set_yticks(range(len(all_cats)))
            ax.set_yticklabels(['Cat %d' % c for c in all_cats])

            for i in range(len(all_cats)):
                for j in range(len(HORIZONS)):
                    v = matrix[i, j]
                    if v > 0.001:
                        tc = 'white' if v > 0.5 else 'black'
                        ax.text(j, i, '%.2f' % v, ha='center',
                                va='center', fontsize=8,
                                color=tc, fontweight='bold')

            ax.set_title(hurr, fontsize=11, fontweight='bold',
                         color=HCOLORS[hurr])
            ax.set_xlabel('Horizon')
            ax.set_ylabel('Category')

        fig.colorbar(im, ax=axes, label='$P$(category $|$ $\\tau$)',
                     shrink=0.6)
        fig.suptitle('Category probability heatmaps by hurricane',
                     fontsize=13, fontweight='bold', y=1.01)
        fig.subplots_adjust(hspace=0.35, wspace=0.30)
        self._save(fig, 'fig_cat_heatmap.pdf')

    # ═════════════════════════════════════════════════════════════════
    # FIGURA: Evolución Factor Demanda
    # ═════════════════════════════════════════════════════════════════

    def fig_demand_evolution(self):
        """Factor demanda ponderado E[η] y rango por horizonte."""
        fig, ax = plt.subplots(figsize=(8, 5))

        for hurr in HURRICANES:
            wds, lo, hi = [], [], []
            for h in HORIZONS:
                ss = self.cascades[hurr][h]
                wd = sum(s.probability * s.demand_factor
                         for s in ss.scenarios)
                ds = [s.demand_factor for s in ss.scenarios]
                wds.append(wd)
                lo.append(wd - min(ds))
                hi.append(max(ds) - wd)

            ax.errorbar(range(len(HORIZONS)), wds,
                        yerr=[lo, hi], fmt='-' + HMARKERS[hurr],
                        color=HCOLORS[hurr], linewidth=2,
                        markersize=8, capsize=4,
                        label=hurr)

        ax.set_xticks(range(len(HORIZONS)))
        ax.set_xticklabels(['%dh' % h for h in HORIZONS])
        ax.set_xlabel('Forecast horizon $\\tau$')
        ax.set_ylabel('Demand factor $\\eta$')
        ax.set_title('Weighted demand factor $E[\\eta]$ '
                     'with scenario range')
        ax.legend(fontsize=9)
        ax.axhline(y=1.0, color='gray', linestyle=':',
                   linewidth=1, alpha=0.5)

        fig.tight_layout()
        self._save(fig, 'fig_demand_evolution.pdf')

    # ═════════════════════════════════════════════════════════════════
    # EXCEL ENRIQUECIDO
    # ═════════════════════════════════════════════════════════════════

    def export_excel(self, filepath: str = "escenarios_validacion.xlsx"):
        """Exporta Excel con 10 hojas de análisis."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("  ⚠ openpyxl no disponible, saltando Excel")
            return

        wb = openpyxl.Workbook()
        hfont = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill("solid", fgColor="1A5276")
        halign = Alignment(horizontal="center", vertical="center",
                           wrap_text=True)
        brd = Border(left=Side("thin"), right=Side("thin"),
                     top=Side("thin"), bottom=Side("thin"))

        def hdr(ws, headers, row=1):
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font = hfont
                cell.fill = hfill
                cell.alignment = halign
                cell.border = brd
                ws.column_dimensions[
                    get_column_letter(c)].width = max(14, len(h) + 4)

        # ─── Hoja 1: Resumen ───────────────────────────────────
        ws = wb.active
        ws.title = "Resumen"
        ws["A1"].value = "MÓDULO EXPERIMENTAL — VALIDACIÓN ESCENARIOS"
        ws["A1"].font = Font(bold=True, size=14, color="1A5276")
        rows_data = [
            ("Fecha:", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Huracanes:", ", ".join(HURRICANES)),
            ("Horizontes:", ", ".join("%dh" % h for h in HORIZONS)),
            ("Método:", "GMM (EM) + Bayesian Updating"),
            ("Seed:", str(SEED)),
            ("Bootstrap B:", str(N_BOOTSTRAP)),
            ("Base datos:", "%d tormentas" % self.gen.n_storms),
        ]
        for i, (l, v) in enumerate(rows_data, 3):
            ws.cell(row=i, column=1, value=l).font = Font(bold=True)
            ws.cell(row=i, column=2, value=v)
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 50

        # ─── Hoja 2: Escenarios completos ─────────────────────
        ws2 = wb.create_sheet("Escenarios")
        hdr(ws2, ["Huracán", "Horizonte(h)", "ω_s", "Etiqueta",
                  "P(ω_s)", "Categoría", "Viento(kt)", "η_s",
                  "P_s(60)", "Lat", "Lon", "Localidad",
                  "Provincia", "Análogos"])
        r = 2
        for hurr in HURRICANES:
            for h in HORIZONS:
                ss = self.cascades[hurr][h]
                for s in ss.scenarios:
                    Ps = int(round(s.demand_factor * 60))
                    vals = [hurr, h, "ω%d" % s.id, s.label,
                            round(s.probability, 6), s.category,
                            s.max_wind_kt,
                            round(s.demand_factor, 4), Ps,
                            round(s.impact_lat, 4),
                            round(s.impact_lon, 4),
                            s.impact_location,
                            s.impact_province,
                            ", ".join(s.reference_storms[:3])]
                    for c, v in enumerate(vals, 1):
                        ws2.cell(row=r, column=c, value=v).border = brd
                    r += 1

        # ─── Hoja 3: Convergencia Bayesiana ───────────────────
        ws3 = wb.create_sheet("Convergencia Bayesiana")
        hdr(ws3, ["Huracán", "Horizonte(h)", "H (bits)",
                  "P(1ro)", "P(2do)", "P(3ro)",
                  "Activos(P>0.01)", "σ²"])
        sigma2 = {72: 1.0, 48: 0.6, 36: 0.4, 24: 0.25, 12: 0.15}
        r = 2
        for hurr in HURRICANES:
            for h in HORIZONS:
                ss = self.cascades[hurr][h]
                ps = sorted([s.probability for s in ss.scenarios],
                            reverse=True)
                ent = self._entropy(ss)
                active = sum(1 for p in ps if p > 0.01)
                vals = [hurr, h, round(ent, 4),
                        round(ps[0], 6) if ps else 0,
                        round(ps[1], 6) if len(ps) > 1 else 0,
                        round(ps[2], 6) if len(ps) > 2 else 0,
                        active, sigma2.get(h, 0.5)]
                for c, v in enumerate(vals, 1):
                    ws3.cell(row=r, column=c, value=v).border = brd
                r += 1

        # ─── Hoja 4: Bootstrap Estabilidad ────────────────────
        ws4 = wb.create_sheet("Bootstrap Estabilidad")
        hdr(ws4, ["Huracán", "ω_s", "p̄_s", "σ(p_s)",
                  "CV(%)", "Rango estable"])
        r = 2
        for hurr in HURRICANES:
            ss72 = self.cascades[hurr][72]
            for s in ss72.scenarios:
                bs_probs = []
                for b in range(N_BOOTSTRAP):
                    bs_ss = self.bootstrap[hurr][72][b]
                    for bs in bs_ss.scenarios:
                        if bs.id == s.id:
                            bs_probs.append(bs.probability)
                            break
                if bs_probs:
                    mean_p = np.mean(bs_probs)
                    std_p = np.std(bs_probs)
                    cv = (std_p / mean_p * 100) if mean_p > 0 else 0
                    rank_ok = "✓"
                else:
                    mean_p, std_p, cv, rank_ok = 0, 0, 0, "✗"

                vals = [hurr, "ω%d" % s.id,
                        round(mean_p, 4), round(std_p, 4),
                        round(cv, 1), rank_ok]
                for c, v in enumerate(vals, 1):
                    ws4.cell(row=r, column=c, value=v).border = brd
                r += 1

        # ─── Hoja 5: Entropía Cruzada ────────────────────────
        ws5 = wb.create_sheet("Entropía Cruzada")
        hdr(ws5, ["Huracán"] + ["%dh" % h for h in HORIZONS]
            + ["Reducción(%)"])
        r = 2
        for hurr in HURRICANES:
            ents = [self._entropy(self.cascades[hurr][h])
                    for h in HORIZONS]
            red = (1 - ents[-1] / ents[0]) * 100 if ents[0] > 0 else 0
            vals = [hurr] + [round(e, 4) for e in ents] + [
                round(red, 1)]
            for c, v in enumerate(vals, 1):
                ws5.cell(row=r, column=c, value=v).border = brd
            r += 1
        # Mean row
        r += 1
        ws5.cell(row=r, column=1, value="PROMEDIO").font = Font(bold=True)
        for j, h in enumerate(HORIZONS):
            mean_e = np.mean([
                self._entropy(self.cascades[hurr][h])
                for hurr in HURRICANES])
            ws5.cell(row=r, column=j + 2,
                     value=round(mean_e, 4)).font = Font(bold=True)

        # ─── Hoja 6: Demanda por Escenario ───────────────────
        ws6 = wb.create_sheet("Demanda")
        hdr(ws6, ["Huracán", "Horizonte", "E[η]", "min(η)",
                  "max(η)", "P_ev (P_nom=60)", "Σπ=75",
                  "Factible P^EV"])
        r = 2
        for hurr in HURRICANES:
            for h in HORIZONS:
                ss = self.cascades[hurr][h]
                etas = [s.demand_factor for s in ss.scenarios]
                wd = sum(s.probability * s.demand_factor
                         for s in ss.scenarios)
                P_ev = int(round(wd * 60))
                feas = "✓" if P_ev <= 75 else "✗"
                vals = [hurr, h, round(wd, 4),
                        round(min(etas), 4),
                        round(max(etas), 4), P_ev, 75, feas]
                for c, v in enumerate(vals, 1):
                    ws6.cell(row=r, column=c, value=v).border = brd
                r += 1

        # ─── Hoja 7: Coordenadas Impacto ─────────────────────
        ws7 = wb.create_sheet("Coordenadas Impacto")
        hdr(ws7, ["Huracán", "Horizonte", "ω_s", "Lat",
                  "Lon", "Localidad", "Provincia",
                  "P(ω_s)", "Categoría"])
        r = 2
        for hurr in HURRICANES:
            for h in HORIZONS:
                ss = self.cascades[hurr][h]
                for s in ss.scenarios:
                    vals = [hurr, h, "ω%d" % s.id,
                            round(s.impact_lat, 4),
                            round(s.impact_lon, 4),
                            s.impact_location,
                            s.impact_province,
                            round(s.probability, 6),
                            s.category]
                    for c, v in enumerate(vals, 1):
                        ws7.cell(row=r, column=c,
                                 value=v).border = brd
                    r += 1

        # ─── Hoja 8: Métricas ML ─────────────────────────────
        ws8 = wb.create_sheet("Métricas ML")
        hdr(ws8, ["Huracán", "Horizonte", "S*",
                  "Silhouette", "BIC", "Método"])
        r = 2
        for hurr in HURRICANES:
            for h in HORIZONS:
                ss = self.cascades[hurr][h]
                vals = [hurr, h, ss.n_scenarios,
                        round(ss.silhouette_score, 4)
                        if ss.silhouette_score else "N/A",
                        round(ss.bic, 2) if ss.bic else "N/A",
                        ss.clustering_method]
                for c, v in enumerate(vals, 1):
                    ws8.cell(row=r, column=c,
                             value=v).border = brd
                r += 1

        # ─── Hoja 9: Análogos (Ian 72h) ─────────────────────
        ws9 = wb.create_sheet("Análogos Ian")
        ss = self.cascades['IAN'][72]
        hdr(ws9, ["ω_s", "Análogos históricos", "Tamaño cluster"])
        r = 2
        for s in ss.scenarios:
            vals = ["ω%d" % s.id,
                    ", ".join(s.reference_storms),
                    s.parameters.get('cluster_size', '')]
            for c, v in enumerate(vals, 1):
                ws9.cell(row=r, column=c, value=v).border = brd
            r += 1

        # ─── Hoja 10: Metadatos ──────────────────────────────
        ws10 = wb.create_sheet("Metadatos")
        meta = [
            ("Módulo:", "ModuloExperimentalValidacion.py"),
            ("Motor:", "GeneradorEscenariosHuracan.py v2.0"),
            ("", ""),
            ("Definiciones formales:", ""),
            ("P^EV:",
             "Problema determinista con demanda media "
             "d̄ = Σ p_s·d_s"),
            ("EEV:",
             "Evaluar solución P^EV bajo todos los escenarios"),
            ("VSS:",
             "Valor de la Solución Estocástica = EEV − Z*_{P^S}"),
            ("H(Ξ):",
             "Entropía de Shannon = −Σ p_s·log₂(p_s)"),
            ("η_s:",
             "Factor demanda = 0.6 + 0.15·cat + 0.005·transit"),
            ("P_s:",
             "Demanda escenario = η_s × P_nom"),
            ("Σπ:",
             "Capacidad total refugios"),
            ("σ²(τ):",
             "Precisión Bayesiana: "
             "{72:1.0, 48:0.6, 36:0.4, 24:0.25, 12:0.15}"),
            ("", ""),
            ("Referencia:", "Birge & Louveaux (2011), "
             "Introduction to Stochastic Programming"),
        ]
        for i, (l, v) in enumerate(meta, 1):
            ws10.cell(row=i, column=1, value=l).font = Font(bold=True)
            ws10.cell(row=i, column=2, value=v)
        ws10.column_dimensions["A"].width = 25
        ws10.column_dimensions["B"].width = 65

        wb.save(filepath)
        print(f"  ✓ {filepath} ({wb.sheetnames})")

    # ═════════════════════════════════════════════════════════════════
    # EJECUCIÓN COMPLETA
    # ═════════════════════════════════════════════════════════════════

    def run_all(self):
        """Ejecuta generación completa: datos + 8 figuras + Excel."""
        t0 = time.time()

        self.generate_all_data()

        print("═" * 60)
        print("  GENERANDO FIGURAS DE PUBLICACIÓN")
        print("═" * 60)

        self.fig_framework()
        self.fig_silhouette_bic()
        self.fig_prob_convergence()
        self.fig_impact_map()
        self.fig_uncertainty_ab()
        self.fig_entropy_multi()
        self.fig_cat_heatmap()
        self.fig_demand_evolution()

        print("\n" + "═" * 60)
        print("  EXPORTANDO EXCEL ENRIQUECIDO")
        print("═" * 60)

        xlsx_path = os.path.join(self.output_dir,
                                 "escenarios_validacion.xlsx")
        self.export_excel(xlsx_path)

        elapsed = time.time() - t0
        print("\n" + "═" * 60)
        print("  COMPLETADO en %.1f s" % elapsed)
        print("═" * 60)
        print(f"\n  📁 Figuras: {self.output_dir}/")
        print(f"  📊 Excel:   {xlsx_path}")
        print(f"  📄 8 PDFs generados\n")


# ═════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    gfp = GeneradorFigurasPublicacion(output_dir="pub_figures")
    gfp.run_all()
