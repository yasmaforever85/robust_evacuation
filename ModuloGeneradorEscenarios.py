#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Modulo V: Generador de Escenarios de Huracan - Interfaz Grafica v2.1
Compatible con GeneradorEscenariosHuracan.py v2.0

Pestanas:
  1. Analisis Historico (escenarios + coordenadas de impacto)
  2. Cascada Temporal (convergencia bayesiana 72-12h)
  3. Analisis y Reportes (graficos + exportacion Excel)
  4. Tiempo Real NHC
  5. Catalogo

Dependencias: PyQt6, numpy, scipy, matplotlib, openpyxl
Autor: Y. Fernandez-Fernandez - UPEC
Version: 2.1 - Febrero 2026
"""

import sys
import os
import json
import math
from datetime import datetime
from typing import Optional, List, Dict, Any, Tuple

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QTabWidget, QGroupBox, QLabel, QComboBox, QPushButton,
    QSpinBox, QTableWidget, QTableWidgetItem, QTextEdit, QProgressBar,
    QSplitter, QFrame, QHeaderView, QFileDialog, QMessageBox,
    QSizePolicy, QStatusBar, QScrollArea
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import (
    QFont, QColor, QPainter, QPen, QBrush,
    QAction, QPalette, QPixmap, QImage
)

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
import numpy as np

try:
    from GeneradorEscenariosHuracan import (
        HurricaneScenarioEngine, ScenarioSet, Scenario, CubaImpact,
        CubaRegion, HORIZONS, category_label, wind_to_category,
        haversine_km, min_distance_to_cuba, CUBA_PROVINCES, MLEngine,
        DatabaseManager, nearest_locality, CUBA_COASTLINE_POINTS
    )
    ENGINE_OK = True
except ImportError as e:
    ENGINE_OK = False
    ENGINE_ERROR = str(e)

try:
    import openpyxl
    from openpyxl.styles import Font as XlFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# =============================================================================
# COLORES
# =============================================================================

COLORS = {
    "bg": "#F0F4F8", "panel": "#FFFFFF", "primary": "#1A5276",
    "primary_light": "#2980B9", "accent": "#E74C3C", "success": "#27AE60",
    "warning": "#F39C12", "text": "#2C3E50", "text_light": "#7F8C8D",
    "border": "#BDC3C7",
}

CAT_COLORS = {
    -1: "#C0C0C0", 0: "#00CED1", 1: "#FFD700",
     2: "#FFA500",  3: "#FF4500",  4: "#FF0000", 5: "#8B008B",
}

HORIZON_COLORS = {72: "#3498DB", 48: "#2ECC71", 36: "#F1C40F", 24: "#E67E22", 12: "#E74C3C"}

STYLESHEET = """
QMainWindow { background-color: #F0F4F8; }
QGroupBox { font-weight: bold; border: 1px solid #BDC3C7; border-radius: 6px;
    margin-top: 12px; padding-top: 14px; background-color: white; }
QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 6px; color: #1A5276; }
QPushButton { background-color: #1A5276; color: white; border: none;
    border-radius: 4px; padding: 8px 16px; font-weight: bold; min-height: 28px; }
QPushButton:hover { background-color: #2980B9; }
QPushButton:disabled { background-color: #BDC3C7; }
QPushButton#btnDanger { background-color: #E74C3C; }
QPushButton#btnSuccess { background-color: #27AE60; }
QPushButton#btnWarning { background-color: #F39C12; }
QComboBox, QSpinBox { border: 1px solid #BDC3C7; border-radius: 4px;
    padding: 4px 8px; min-height: 24px; background-color: white; }
QTableWidget { border: 1px solid #BDC3C7; border-radius: 4px;
    gridline-color: #ECF0F1; background-color: white; }
QHeaderView::section { background-color: #1A5276; color: white;
    padding: 6px; border: none; font-weight: bold; }
QTabBar::tab { background-color: #BDC3C7; color: #2C3E50;
    padding: 8px 20px; border-top-left-radius: 4px; border-top-right-radius: 4px; margin-right: 2px; }
QTabBar::tab:selected { background-color: #1A5276; color: white; }
QTextEdit { border: 1px solid #BDC3C7; border-radius: 4px;
    font-family: Consolas, monospace; font-size: 11px; background-color: #FAFAFA; }
QProgressBar { border: 1px solid #BDC3C7; border-radius: 4px;
    text-align: center; background-color: #ECF0F1; }
QProgressBar::chunk { background-color: #2980B9; border-radius: 3px; }
QLabel#statLabel { color: #2C3E50; font-size: 18px; font-weight: bold; }
"""

# =============================================================================
# UTILIDADES GRAFICAS
# =============================================================================

def fig_to_pixmap(fig, dpi=120):
    canvas = FigureCanvasAgg(fig)
    canvas.draw()
    buf = canvas.buffer_rgba()
    w, h = canvas.get_width_height()
    img = QImage(buf, w, h, QImage.Format.Format_RGBA8888)
    pix = QPixmap.fromImage(img)
    plt.close(fig)
    return pix

def pub_style(ax, title="", xlabel="", ylabel="", grid=True):
    ax.set_title(title, fontsize=11, fontweight="bold", color="#1A5276", pad=10)
    ax.set_xlabel(xlabel, fontsize=9, color="#2C3E50")
    ax.set_ylabel(ylabel, fontsize=9, color="#2C3E50")
    ax.tick_params(labelsize=8, colors="#555")
    if grid:
        ax.grid(True, alpha=0.3, linestyle="--")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

# =============================================================================
# WIDGETS PERSONALIZADOS
# =============================================================================

class ProbabilityBar(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumHeight(40); self.setMaximumHeight(50)
        self._segments = []

    def set_data(self, scenarios):
        self._segments = []
        for s in scenarios:
            color = CAT_COLORS.get(s.category, "#1A5276")
            label = "w%d (%.1f%%)" % (s.id, s.probability * 100)
            self._segments.append((s.probability, color, label))
        self.update()

    def paintEvent(self, event):
        if not self._segments: return
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        w, h = self.width() - 4, self.height() - 8
        x = 2
        for prob, color, label in self._segments:
            seg_w = max(2, int(w * prob))
            painter.setBrush(QBrush(QColor(color)))
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawRoundedRect(int(x), 4, seg_w, h, 3, 3)
            if seg_w > 50:
                painter.setPen(QPen(QColor("white")))
                painter.setFont(QFont("Arial", 8, QFont.Weight.Bold))
                painter.drawText(int(x)+4, 4, seg_w-8, h, Qt.AlignmentFlag.AlignVCenter, label)
            x += seg_w + 1
        painter.end()


class ChartLabel(QLabel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setMinimumSize(380, 280)
        self.setStyleSheet("border: 1px solid #BDC3C7; border-radius: 4px; background: white;")

    def set_figure(self, fig, dpi=120):
        pixmap = fig_to_pixmap(fig, dpi)
        scaled = pixmap.scaled(self.size(), Qt.AspectRatioMode.KeepAspectRatio,
                               Qt.TransformationMode.SmoothTransformation)
        self.setPixmap(scaled)

# =============================================================================
# HILO DE PROCESAMIENTO
# =============================================================================

class WorkerThread(QThread):
    progress = pyqtSignal(str, int)
    finished = pyqtSignal(object)
    error = pyqtSignal(str)

    def __init__(self, engine, task, params):
        super().__init__()
        self.engine = engine; self.task = task; self.params = params

    def run(self):
        try:
            if self.task == "analyze":
                self.progress.emit("Buscando analogos...", 20)
                result = self.engine.analyze(**self.params)
                self.progress.emit("Listo", 100)
                self.finished.emit(result)

            elif self.task == "cascade":
                self.progress.emit("Generando cascada...", 10)
                result = self.engine.cascade(**self.params)
                self.progress.emit("Listo", 100)
                self.finished.emit(result)

            elif self.task == "full_analysis":
                name = self.params["hurricane_name"]
                region = self.params.get("region")

                self.progress.emit("Cascada temporal...", 15)
                cascade = self.engine.cascade(name, region=region, method="gmm")

                self.progress.emit("Comparando metodos...", 40)
                compare = {}
                for m in ["gmm", "kmeans"]:
                    compare[m] = self.engine.analyze(name, horizon=48, region=region, method=m)

                self.progress.emit("Seleccion optima k...", 60)
                target_impact = None
                for imp in self.engine.generator.impacts:
                    if imp.track.name.upper() == name.upper():
                        target_impact = imp; break

                analogs = self.engine.generator.find_analogs(name, n_analogs=18)
                af = np.array([a[0].feature_vector for a in analogs])
                k_sil, k_bic = {}, {}
                for k in range(2, min(8, len(af))):
                    lk, _, _ = MLEngine.kmeans(af, k)
                    k_sil[k] = MLEngine.silhouette_score(af, lk)
                    _, _, _, _, _, bk = MLEngine.gmm(af, k)
                    k_bic[k] = bk

                self.progress.emit("Detalles de analogos...", 80)
                analog_details = []
                for imp_a, sim in analogs:
                    analog_details.append({
                        "name": imp_a.track.name, "year": imp_a.track.year,
                        "category": imp_a.category_at_closest,
                        "wind_kt": imp_a.wind_at_closest,
                        "lat": round(imp_a.closest_point.lat, 2),
                        "lon": round(imp_a.closest_point.lon, 2),
                        "region": imp_a.region.value, "landfall": imp_a.landfall,
                        "distance_km": round(imp_a.closest_approach_km, 1),
                        "transit_h": round(imp_a.transit_hours, 1),
                        "bearing": round(imp_a.entry_bearing, 1),
                        "similarity": round(sim, 4),
                        "provinces": imp_a.provinces_affected,
                    })

                self.progress.emit("Listo", 100)
                self.finished.emit({
                    "type": "full_analysis", "cascade": cascade,
                    "compare": compare, "k_silhouettes": k_sil, "k_bics": k_bic,
                    "analog_details": analog_details, "target": target_impact,
                })

            elif self.task == "realtime":
                self.progress.emit("Consultando NHC...", 30)
                result = self.engine.check_active_threats()
                self.progress.emit("Listo", 100)
                self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))

# =============================================================================
# GENERADOR DE GRAFICOS
# =============================================================================

class Charts:
    @staticmethod
    def prob_dist(ss):
        fig, ax = plt.subplots(figsize=(5, 3.5), dpi=120)
        ids = ["$\\omega_{%d}$" % s.id for s in ss.scenarios]
        probs = [s.probability for s in ss.scenarios]
        colors = [CAT_COLORS.get(s.category, "#1A5276") for s in ss.scenarios]
        bars = ax.bar(ids, probs, color=colors, edgecolor="white", linewidth=0.8)
        for bar, p in zip(bars, probs):
            if p > 0.01:
                ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.01,
                        "%.3f" % p, ha="center", va="bottom", fontsize=8, fontweight="bold")
        pub_style(ax, "Distribucion de Probabilidad", "Escenario", "P(omega)")
        ax.set_ylim(0, max(probs)*1.2 if probs else 1); fig.tight_layout()
        return fig

    @staticmethod
    def bayesian(cascade):
        fig, ax = plt.subplots(figsize=(5.5, 3.8), dpi=120)
        sh = sorted(cascade.keys(), reverse=True)
        hlabels = ["%dh" % h for h in sh]
        if not sh: return fig
        scenario_ids = sorted(set(s.id for s in cascade[sh[0]].scenarios))
        lc = ["#E74C3C","#2980B9","#27AE60","#F39C12","#8B008B","#00CED1"]
        mk = ["o","s","^","D","v","P"]
        for idx, sid in enumerate(scenario_ids):
            probs = []
            for h in sh:
                p = 0.0
                for s in cascade[h].scenarios:
                    if s.id == sid: p = s.probability; break
                probs.append(p)
            c = lc[idx % len(lc)]; m = mk[idx % len(mk)]
            ax.plot(hlabels, probs, "-"+m, color=c, linewidth=2, markersize=7,
                    label="$\\omega_{%d}$" % sid)
        pub_style(ax, "Convergencia Bayesiana", "Horizonte", "P(omega|h)")
        ax.legend(fontsize=8, loc="upper left", framealpha=0.9)
        ax.set_ylim(-0.02, 1.05); fig.tight_layout()
        return fig

    @staticmethod
    def entropy(cascade):
        fig, ax = plt.subplots(figsize=(5, 3.5), dpi=120)
        sh = sorted(cascade.keys(), reverse=True)
        ents = []
        for h in sh:
            e = -sum(s.probability*math.log2(s.probability)
                     for s in cascade[h].scenarios if s.probability > 1e-15)
            ents.append(e)
        labels = ["%dh" % h for h in sh]
        colors = [HORIZON_COLORS.get(h, "#1A5276") for h in sh]
        ax.bar(labels, ents, color=colors, edgecolor="white")
        ax.plot(labels, ents, "k--o", markersize=5, linewidth=1.5, alpha=0.7)
        for i, e in enumerate(ents):
            ax.text(i, e+0.03, "%.3f" % e, ha="center", fontsize=8, fontweight="bold")
        pub_style(ax, "Entropia de Shannon H(Xi)", "Horizonte", "H (bits)")
        ax.set_ylim(0, max(ents)*1.25 if ents else 2); fig.tight_layout()
        return fig

    @staticmethod
    def sil_k(k_sil):
        fig, ax = plt.subplots(figsize=(4.5, 3.2), dpi=120)
        ks = sorted(k_sil.keys()); scores = [k_sil[k] for k in ks]
        ax.plot(ks, scores, "-o", color="#2980B9", linewidth=2, markersize=8)
        bk = max(k_sil, key=k_sil.get)
        ax.axvline(bk, color="#E74C3C", linestyle="--", alpha=0.7,
                   label="k*=%d (S=%.3f)" % (bk, k_sil[bk]))
        for k, s in zip(ks, scores):
            ax.text(k, s+0.01, "%.3f" % s, ha="center", fontsize=8)
        pub_style(ax, "Silhouette Score vs k", "k", "Silhouette")
        ax.legend(fontsize=9); ax.set_xticks(ks); fig.tight_layout()
        return fig

    @staticmethod
    def bic_k(k_bic):
        fig, ax = plt.subplots(figsize=(4.5, 3.2), dpi=120)
        ks = sorted(k_bic.keys()); bics = [k_bic[k] for k in ks]
        ax.plot(ks, bics, "-s", color="#8B008B", linewidth=2, markersize=8)
        bk = min(k_bic, key=k_bic.get)
        ax.axvline(bk, color="#27AE60", linestyle="--", alpha=0.7,
                   label="k*=%d (BIC=%.1f)" % (bk, k_bic[bk]))
        pub_style(ax, "BIC vs k (GMM)", "k", "BIC")
        ax.legend(fontsize=9); ax.set_xticks(ks); fig.tight_layout()
        return fig

    @staticmethod
    def method_cmp(compare):
        fig, axes = plt.subplots(1, 2, figsize=(8, 3.5), dpi=120)
        for idx, (method, ss) in enumerate(compare.items()):
            ax = axes[idx]
            ids = ["$\\omega_{%d}$" % s.id for s in ss.scenarios]
            probs = [s.probability for s in ss.scenarios]
            colors = [CAT_COLORS.get(s.category, "#1A5276") for s in ss.scenarios]
            ax.bar(ids, probs, color=colors, edgecolor="white")
            ml = "GMM (EM)" if method == "gmm" else "K-Means++"
            t = "%s (S=%.3f)" % (ml, ss.silhouette_score)
            if ss.bic is not None: t += " BIC=%.1f" % ss.bic
            pub_style(ax, t, "Escenario", "P(omega)")
            ax.set_ylim(0, max(probs)*1.25 if probs else 1)
        fig.suptitle("Comparacion de Metodos (48h)", fontsize=11,
                     fontweight="bold", color="#1A5276", y=1.02)
        fig.tight_layout(); return fig

    @staticmethod
    def cat_heat(cascade):
        fig, ax = plt.subplots(figsize=(5.5, 3.5), dpi=120)
        sh = sorted(cascade.keys(), reverse=True)
        all_cats = sorted(set(s.category for ss in cascade.values() for s in ss.scenarios))
        if not all_cats or not sh: return fig
        matrix = np.zeros((len(all_cats), len(sh)))
        for j, h in enumerate(sh):
            for s in cascade[h].scenarios:
                i = all_cats.index(s.category)
                matrix[i, j] += s.probability
        im = ax.imshow(matrix, aspect="auto", cmap="YlOrRd", interpolation="nearest")
        ax.set_xticks(range(len(sh))); ax.set_xticklabels(["%dh" % h for h in sh])
        ax.set_yticks(range(len(all_cats)))
        ax.set_yticklabels([category_label(c) for c in all_cats])
        for i in range(len(all_cats)):
            for j in range(len(sh)):
                v = matrix[i,j]
                if v > 0.001:
                    tc = "white" if v > 0.5 else "black"
                    ax.text(j, i, "%.2f" % v, ha="center", va="center", fontsize=8,
                            color=tc, fontweight="bold")
        fig.colorbar(im, ax=ax, label="P(cat|h)", shrink=0.8)
        pub_style(ax, "Probabilidad por Categoria y Horizonte", "Horizonte",
                  "Categoria", grid=False)
        fig.tight_layout(); return fig

    @staticmethod
    def demand(cascade):
        fig, ax = plt.subplots(figsize=(5, 3.5), dpi=120)
        sh = sorted(cascade.keys(), reverse=True)
        wds = []
        ranges = []
        for h in sh:
            ss = cascade[h]
            wd = sum(s.probability*s.demand_factor for s in ss.scenarios)
            wds.append(wd)
            ds = [s.demand_factor for s in ss.scenarios]
            ranges.append((min(ds), max(ds)))
        labels = ["%dh" % h for h in sh]
        colors = [HORIZON_COLORS.get(h, "#1A5276") for h in sh]
        ax.bar(labels, wds, color=colors, edgecolor="white", alpha=0.8)
        lows = [wd - r[0] for wd, r in zip(wds, ranges)]
        highs = [r[1] - wd for wd, r in zip(wds, ranges)]
        ax.errorbar(labels, wds, yerr=[lows, highs], fmt="none",
                    ecolor="black", capsize=4, linewidth=1.5)
        for i, wd in enumerate(wds):
            ax.text(i, wd+0.02, "%.3f" % wd, ha="center", fontsize=8, fontweight="bold")
        pub_style(ax, "Factor Demanda Ponderado E[D_omega]", "Horizonte", "E[D]")
        fig.tight_layout(); return fig

    @staticmethod
    def impact_map(ss, cuba_coast=None):
        """Scatter map of scenario impact points with Cuba outline."""
        fig, ax = plt.subplots(figsize=(6, 3.8), dpi=120)
        # Draw Cuba coastline approximation
        if cuba_coast:
            # Split into north and south coast
            nc = [(lat, lon) for lat, lon in cuba_coast if lat > 21.3 or lon > -83]
            sc = [(lat, lon) for lat, lon in cuba_coast if lat <= 21.3 or lon <= -83]
            nc.sort(key=lambda p: p[1])
            sc.sort(key=lambda p: p[1])
            if nc:
                ax.plot([p[1] for p in nc], [p[0] for p in nc],
                        '-', color='#A0A0A0', linewidth=1.5, alpha=0.5)
            if sc:
                ax.plot([p[1] for p in sc], [p[0] for p in sc],
                        '-', color='#A0A0A0', linewidth=1.5, alpha=0.5)
            # Fill between to shade Cuba
            all_pts = sorted(cuba_coast, key=lambda p: p[1])
            ax.fill([p[1] for p in all_pts], [p[0] for p in all_pts],
                    alpha=0.08, color='#2C3E50')

        for s in ss.scenarios:
            if s.impact_lat == 0 and s.impact_lon == 0:
                continue
            size = max(80, s.probability * 2000)
            color = CAT_COLORS.get(s.category, "#1A5276")
            ax.scatter(s.impact_lon, s.impact_lat, s=size, c=color,
                       edgecolors='black', linewidth=0.8, alpha=0.85, zorder=5)
            label_txt = "$\\omega_{%d}$\n%s\n%.1f%%" % (
                s.id, s.impact_location, s.probability * 100)
            ax.annotate(label_txt, (s.impact_lon, s.impact_lat),
                        textcoords="offset points", xytext=(12, 8),
                        fontsize=7, fontweight="bold",
                        bbox=dict(boxstyle="round,pad=0.2", fc="white", alpha=0.85),
                        arrowprops=dict(arrowstyle="-", color="gray", alpha=0.5))

        pub_style(ax, "Puntos de Impacto Probables por Escenario (%dh)" % ss.horizon_hours,
                  "Longitud", "Latitud", grid=True)
        ax.set_xlim(-86, -73.5); ax.set_ylim(19.5, 24)
        ax.set_aspect('auto')
        fig.tight_layout(); return fig

    @staticmethod
    def impact_cascade(cascade, cuba_coast=None):
        """Impact location evolution across horizons for most probable scenario."""
        fig, ax = plt.subplots(figsize=(6, 3.8), dpi=120)
        if cuba_coast:
            all_pts = sorted(cuba_coast, key=lambda p: p[1])
            ax.fill([p[1] for p in all_pts], [p[0] for p in all_pts],
                    alpha=0.08, color='#2C3E50')
            nc = sorted(cuba_coast, key=lambda p: p[1])
            ax.plot([p[1] for p in nc], [p[0] for p in nc],
                    '-', color='#A0A0A0', linewidth=1, alpha=0.4)

        sh = sorted(cascade.keys(), reverse=True)
        for h in sh:
            ss = cascade[h]
            best = max(ss.scenarios, key=lambda s: s.probability)
            if best.impact_lat == 0: continue
            color = HORIZON_COLORS.get(h, "#1A5276")
            size = max(60, best.probability * 800)
            ax.scatter(best.impact_lon, best.impact_lat, s=size, c=color,
                       edgecolors='black', linewidth=0.8, alpha=0.85, zorder=5)
            ax.annotate("%dh\n%s\nP=%.2f" % (h, best.impact_location, best.probability),
                        (best.impact_lon, best.impact_lat),
                        textcoords="offset points", xytext=(10, 6),
                        fontsize=7, fontweight="bold",
                        bbox=dict(boxstyle="round,pad=0.2", fc="white", alpha=0.85))

        pub_style(ax, "Evolucion del Punto de Impacto Mas Probable",
                  "Longitud", "Latitud", grid=True)
        ax.set_xlim(-86, -73.5); ax.set_ylim(19.5, 24)
        fig.tight_layout(); return fig

# =============================================================================
# EXPORTADOR EXCEL
# =============================================================================

class ExcelExporter:
    @staticmethod
    def export(filepath, data):
        if not EXCEL_OK: raise RuntimeError("openpyxl no disponible")
        wb = openpyxl.Workbook()
        hfont = XlFont(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill("solid", fgColor="1A5276")
        halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
        brd = Border(left=Side("thin"), right=Side("thin"),
                     top=Side("thin"), bottom=Side("thin"))

        def hdr(ws, headers, row=1):
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font = hfont; cell.fill = hfill; cell.alignment = halign; cell.border = brd
                ws.column_dimensions[get_column_letter(c)].width = max(14, len(h)+4)

        # 1. Resumen
        ws = wb.active; ws.title = "Resumen"
        ws.merge_cells("A1:F1")
        ws["A1"].value = "HURRICANE SCENARIO ENGINE v2.0 - REPORTE"
        ws["A1"].font = XlFont(bold=True, size=14, color="1A5276")
        target = data.get("target")
        cascade = data.get("cascade", {})
        rows_data = [
            ("Fecha:", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Huracan:", target.track.name if target else "N/A"),
            ("Ano:", str(target.track.year) if target else ""),
            ("Region:", target.region.value if target else ""),
            ("Categoria:", str(target.category_at_closest) if target else ""),
            ("Viento:", "%d kt" % target.wind_at_closest if target else ""),
            ("Lat impacto:", "%.4f" % target.closest_point.lat if target else ""),
            ("Lon impacto:", "%.4f" % target.closest_point.lon if target else ""),
            ("Landfall:", "Si" if (target and target.landfall) else "No"),
            ("Dist costa:", "%.1f km" % target.closest_approach_km if target else ""),
            ("Provincias:", ", ".join(target.provinces_affected[:5]) if target else ""),
            ("Analogos:", str(len(data.get("analog_details", [])))),
            ("Horizontes:", ", ".join("%dh" % h for h in sorted(cascade.keys()))),
            ("Metodo:", "GMM (EM) + K-Means++ comparison"),
            ("Modelo pronostico:", "HURDAT2 Best Track (NHC)"),
        ]
        for i, (l, v) in enumerate(rows_data, 3):
            ws.cell(row=i, column=1, value=l).font = XlFont(bold=True)
            ws.cell(row=i, column=2, value=v)
        ws.column_dimensions["A"].width = 20; ws.column_dimensions["B"].width = 50

        # 2. Escenarios
        ws2 = wb.create_sheet("Escenarios por Horizonte")
        hdr(ws2, ["Horizonte(h)", "Escenario", "Etiqueta", "P(omega)", "Categoria",
                  "Viento(kt)", "D_omega", "Lat.Impacto", "Lon.Impacto",
                  "Lugar.Impacto", "Provincia", "Analogos", "Tam.cluster",
                  "Cat.prom", "Viento.prom", "Transit.prom(h)", "Ruido.horizonte"])
        r = 2
        for h in sorted(cascade.keys(), reverse=True):
            for s in cascade[h].scenarios:
                p = s.parameters
                vals = [h, "w%d"%s.id, s.label, round(s.probability,6), s.category,
                        s.max_wind_kt, round(s.demand_factor,4),
                        round(s.impact_lat, 4), round(s.impact_lon, 4),
                        s.impact_location, s.impact_province,
                        ", ".join(s.reference_storms[:4]),
                        p.get("cluster_size",""), p.get("avg_category",""),
                        p.get("avg_wind_kt",""), p.get("avg_transit_h",""),
                        p.get("horizon_noise","")]
                for c, v in enumerate(vals, 1):
                    ws2.cell(row=r, column=c, value=v).border = brd
                r += 1

        # 3. Convergencia
        ws3 = wb.create_sheet("Convergencia Bayesiana")
        prec = {72: 1.0, 48: 0.6, 36: 0.4, 24: 0.25, 12: 0.15}
        hdr(ws3, ["Horizonte(h)", "Entropia H(bits)", "P(1ro)", "P(2do)", "P(3ro)",
                  "Escenarios(P>0.01)", "Sigma^2"])
        r = 2
        for h in sorted(cascade.keys(), reverse=True):
            ss = cascade[h]
            ps = sorted([s.probability for s in ss.scenarios], reverse=True)
            ent = -sum(p*math.log2(p) for p in ps if p > 1e-15)
            active = sum(1 for p in ps if p > 0.01)
            vals = [h, round(ent,4),
                    round(ps[0],6) if len(ps)>0 else 0,
                    round(ps[1],6) if len(ps)>1 else 0,
                    round(ps[2],6) if len(ps)>2 else 0,
                    active, prec.get(h, 0.5)]
            for c, v in enumerate(vals, 1):
                ws3.cell(row=r, column=c, value=v).border = brd
            r += 1

        # 4. Comparacion
        ws4 = wb.create_sheet("Comparacion Metodos")
        hdr(ws4, ["Metodo", "N.escenarios", "Silhouette", "BIC",
                  "Escenario.dominante", "P(dominante)", "Entropia"])
        r = 2
        for method, ss in data.get("compare", {}).items():
            ml = "GMM (EM)" if method == "gmm" else "K-Means++"
            best = max(ss.scenarios, key=lambda x: x.probability) if ss.scenarios else None
            ent = -sum(s.probability*math.log2(s.probability)
                       for s in ss.scenarios if s.probability > 1e-15)
            vals = [ml, ss.n_scenarios, round(ss.silhouette_score,4),
                    round(ss.bic,2) if ss.bic else "N/A",
                    "w%d" % best.id if best else "", round(best.probability,4) if best else 0,
                    round(ent,4)]
            for c, v in enumerate(vals, 1):
                ws4.cell(row=r, column=c, value=v).border = brd
            r += 1

        # 5. Seleccion k
        ws5 = wb.create_sheet("Seleccion k")
        hdr(ws5, ["k", "Silhouette", "BIC", "Sil.Optimo", "BIC.Optimo"])
        k_sil = data.get("k_silhouettes", {}); k_bic = data.get("k_bics", {})
        bsk = max(k_sil, key=k_sil.get) if k_sil else None
        bbk = min(k_bic, key=k_bic.get) if k_bic else None
        r = 2
        for k in sorted(set(list(k_sil.keys())+list(k_bic.keys()))):
            vals = [k, round(k_sil.get(k,0),4), round(k_bic.get(k,0),2),
                    "***" if k==bsk else "", "***" if k==bbk else ""]
            for c, v in enumerate(vals, 1):
                ws5.cell(row=r, column=c, value=v).border = brd
            r += 1

        # 6. Analogos
        ws6 = wb.create_sheet("Analogos Historicos")
        hdr(ws6, ["Nombre", "Ano", "Cat", "Viento(kt)", "Lat", "Lon", "Region",
                  "Landfall", "Dist(km)", "Transit(h)", "Rumbo(deg)", "Similitud", "Provincias"])
        r = 2
        for a in data.get("analog_details", []):
            vals = [a["name"], a["year"], a["category"], a["wind_kt"],
                    a["lat"], a["lon"], a["region"],
                    "Si" if a["landfall"] else "No",
                    a["distance_km"], a["transit_h"], a["bearing"],
                    a["similarity"], ", ".join(a["provinces"][:4])]
            for c, v in enumerate(vals, 1):
                ws6.cell(row=r, column=c, value=v).border = brd
            r += 1

        # 7. Coordenadas
        ws7 = wb.create_sheet("Coordenadas Impacto")
        hdr(ws7, ["Huracan", "Ano", "Lat", "Lon", "Lat.Landfall", "Lon.Landfall",
                  "Region", "Cat", "Viento(kt)", "Dist.Costa(km)", "Modelo.Pronostico"])
        r = 2
        if target:
            lf = target.landfall_point
            vals = [target.track.name, target.track.year,
                    round(target.closest_point.lat,4), round(target.closest_point.lon,4),
                    round(lf.lat,4) if lf else "N/A", round(lf.lon,4) if lf else "N/A",
                    target.region.value, target.category_at_closest,
                    target.wind_at_closest, round(target.closest_approach_km,1),
                    "HURDAT2 Best Track (NHC)"]
            for c, v in enumerate(vals, 1):
                cell = ws7.cell(row=r, column=c, value=v)
                cell.border = brd; cell.font = XlFont(bold=True)
                cell.fill = PatternFill("solid", fgColor="D5E8D4")
            r += 1
        for a in data.get("analog_details", []):
            vals = [a["name"], a["year"], a["lat"], a["lon"], "-", "-",
                    a["region"], a["category"], a["wind_kt"], a["distance_km"],
                    "HURDAT2 Best Track (NHC)"]
            for c, v in enumerate(vals, 1):
                ws7.cell(row=r, column=c, value=v).border = brd
            r += 1

        # 8. Metricas ML
        ws8 = wb.create_sheet("Metricas ML")
        ws8["A1"].value = "METRICAS DE MACHINE LEARNING"; ws8["A1"].font = XlFont(bold=True, size=12)
        sections = [
            (3, "1. GMM (EM)", [
                ("Algoritmo:", "Expectation-Maximization (Dempster, Laird & Rubin 1977)"),
                ("Inicializacion:", "K-Means++ warm start"),
                ("Convergencia:", "Delta log-likelihood < 1e-6 o 200 iter"),
                ("Regularizacion:", "Cov + 1e-6*I (estabilidad numerica)"),
                ("Seleccion modelo:", "BIC = -2*ln(L) + p*ln(n)"),
            ]),
            (10, "2. K-Means++", [
                ("Algoritmo:", "Lloyd + K-Means++ init (Arthur & Vassilvitskii 2007)"),
                ("Restarts:", "n_init=10"), ("Convergencia:", "atol=1e-8 o 200 iter"),
                ("Validacion:", "Silhouette Score (Rousseeuw 1987)"),
            ]),
            (16, "3. Bayesian Updating", [
                ("Formula:", "P(omega|h) = P(omega)*L(state|omega,h) / Z"),
                ("Likelihood:", "L = exp(-||state-mu||^2 / (2*sigma^2))"),
                ("sigma^2(72h):", "1.000"), ("sigma^2(48h):", "0.600"),
                ("sigma^2(36h):", "0.400"), ("sigma^2(24h):", "0.250"),
                ("sigma^2(12h):", "0.150"),
            ]),
            (25, "4. Feature Space (9-D)", [
                ("f1:", "Categoria (0-5)"), ("f2:", "Viento norm (wind/185)"),
                ("f3:", "Distancia norm (dist/500)"), ("f4:", "Landfall (0/1)"),
                ("f5:", "Transito norm (h/48)"), ("f6:", "Rumbo norm (bearing/360)"),
                ("f7:", "Region code (Occ=0, Cen=0.5, Ori=1)"),
                ("f8:", "Latitud norm x2 ((lat-19.8)/3.5)*2 [peso geografico]"),
                ("f9:", "Longitud norm x2 ((lon+85)/10.9)*2 [peso geografico]"),
            ]),
            (34, "5. Factor Demanda", [
                ("Formula:", "D = 0.6 + 0.15*avg_cat + 0.005*avg_transit"),
                ("Rango:", "[0.6, ~1.5]"),
                ("Uso:", "Multiplicador demanda evacuacion en PI^S"),
            ]),
        ]
        for start_row, title, params in sections:
            ws8.cell(row=start_row, column=1, value=title).font = XlFont(bold=True, size=11)
            for i, (l, v) in enumerate(params, start_row+1):
                ws8.cell(row=i, column=1, value=l).font = XlFont(bold=True)
                ws8.cell(row=i, column=2, value=v)
        ws8.column_dimensions["A"].width = 22; ws8.column_dimensions["B"].width = 60

        # 9. Impacto por Escenario
        ws9 = wb.create_sheet("Impacto por Escenario")
        hdr(ws9, ["Horizonte(h)", "Escenario", "P(omega)", "Categoria",
                  "Lat.Impacto", "Lon.Impacto", "Lugar.Impacto", "Provincia",
                  "Viento(kt)", "D_omega", "Dist.Localidad(km)", "Analogos"])
        r = 2
        for h in sorted(cascade.keys(), reverse=True):
            for s in cascade[h].scenarios:
                p = s.parameters
                vals = [h, "w%d"%s.id, round(s.probability, 6), s.category,
                        round(s.impact_lat, 4), round(s.impact_lon, 4),
                        s.impact_location, s.impact_province,
                        s.max_wind_kt, round(s.demand_factor, 4),
                        p.get("nearest_locality_km", ""),
                        ", ".join(s.reference_storms[:3])]
                for c, v in enumerate(vals, 1):
                    cell = ws9.cell(row=r, column=c, value=v)
                    cell.border = brd
                # Highlight most probable per horizon
                if s == max(cascade[h].scenarios, key=lambda x: x.probability):
                    for c in range(1, 13):
                        ws9.cell(row=r, column=c).fill = PatternFill("solid", fgColor="D5F5E3")
                r += 1
        ws9.column_dimensions["G"].width = 22; ws9.column_dimensions["H"].width = 20
        ws9.column_dimensions["L"].width = 30

        wb.save(filepath)

# =============================================================================
# VENTANA PRINCIPAL
# =============================================================================

class VentanaPrincipal(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Modulo V - Generador de Escenarios de Huracan v2.1")
        self.setMinimumSize(1200, 800)
        if not ENGINE_OK:
            QMessageBox.critical(self, "Error",
                "No se pudo importar el motor:\n%s" % ENGINE_ERROR)
            sys.exit(1)
        self.engine = HurricaneScenarioEngine()
        self.resultado_actual = None
        self.cascada_actual = None
        self.full_analysis_data = None
        self.worker = None
        self._init_ui()
        self._populate_storms()
        self.statusBar().showMessage(
            "Motor cargado: %d huracanes historicos" % self.engine.database_size)

    def _init_ui(self):
        central = QWidget(); self.setCentralWidget(central)
        self.setStyleSheet(STYLESHEET)
        ml = QVBoxLayout(central); ml.setContentsMargins(8,8,8,8); ml.setSpacing(6)

        # Header
        hf = QFrame()
        hf.setStyleSheet("QFrame{background-color:#1A5276;border-radius:6px;padding:8px;}")
        hl = QHBoxLayout(hf); hl.setContentsMargins(16,8,16,8)
        t = QLabel("  Modulo V: Generador de Escenarios Estocasticos de Huracan")
        t.setStyleSheet("color:white;font-size:16px;font-weight:bold;")
        hl.addWidget(t); hl.addStretch()
        st = QLabel("Cuba | %d huracanes | GMM/K-Means++ | Bayesian | v2.1" % self.engine.database_size)
        st.setStyleSheet("color:#AED6F1;font-size:11px;"); hl.addWidget(st)
        ml.addWidget(hf)

        self.tabs = QTabWidget()
        self.tabs.addTab(self._tab_historico(), " 1. Analisis Historico ")
        self.tabs.addTab(self._tab_cascada(), " 2. Cascada Temporal ")
        self.tabs.addTab(self._tab_analysis(), " 3. Analisis y Reportes ")
        self.tabs.addTab(self._tab_realtime(), " 4. Tiempo Real NHC ")
        self.tabs.addTab(self._tab_catalogo(), " 5. Catalogo ")
        self.tabs.addTab(self._tab_database(), " 6. Base de Datos ")
        ml.addWidget(self.tabs, 1)

        self.progress = QProgressBar(); self.progress.setVisible(False)
        ml.addWidget(self.progress)
        self.setStatusBar(QStatusBar())

        # Menu
        menu = self.menuBar()
        arch = menu.addMenu("Archivo")
        a1 = QAction("Cargar HURDAT2...", self); a1.triggered.connect(self._load_hurdat2); arch.addAction(a1)
        arch.addSeparator()
        a2 = QAction("Salir", self); a2.triggered.connect(self.close); arch.addAction(a2)
        herr = menu.addMenu("Herramientas")
        a3 = QAction("Exportar Excel...", self); a3.triggered.connect(self._export_excel); herr.addAction(a3)
        a4 = QAction("Exportar PI^S JSON...", self); a4.triggered.connect(self._export_pis); herr.addAction(a4)

    # --- TAB 1 ---
    def _tab_historico(self):
        tab = QWidget(); layout = QHBoxLayout(tab)
        left = QGroupBox("Parametros"); left.setMaximumWidth(350)
        ll = QVBoxLayout(left)
        ll.addWidget(QLabel("Huracan:"))
        self.combo_huracan = QComboBox()
        self.combo_huracan.currentIndexChanged.connect(self._on_hchange)
        ll.addWidget(self.combo_huracan)
        self.lbl_info = QLabel("")
        self.lbl_info.setWordWrap(True)
        self.lbl_info.setStyleSheet("background:#F8F9FA;border:1px solid #BDC3C7;border-radius:4px;padding:8px;font-size:11px;")
        ll.addWidget(self.lbl_info)
        ll.addWidget(QLabel("Region:"))
        self.combo_region = QComboBox(); self.combo_region.addItems(["(Auto)","Occidente","Centro","Oriente"])
        ll.addWidget(self.combo_region)
        ll.addWidget(QLabel("Horizonte:"))
        self.combo_h = QComboBox(); self.combo_h.addItems(["72h","48h","36h","24h","12h"]); self.combo_h.setCurrentIndex(1)
        ll.addWidget(self.combo_h)
        ll.addWidget(QLabel("Metodo:"))
        self.combo_met = QComboBox(); self.combo_met.addItems(["GMM (EM)","K-Means++"])
        ll.addWidget(self.combo_met)
        ll.addWidget(QLabel("Escenarios (0=auto):"))
        self.spin_n = QSpinBox(); self.spin_n.setRange(0,8); self.spin_n.setValue(0)
        ll.addWidget(self.spin_n)
        ll.addSpacing(10)
        self.btn_gen = QPushButton("Generar Escenarios"); self.btn_gen.clicked.connect(self._run_analysis)
        ll.addWidget(self.btn_gen)
        self.btn_exp = QPushButton("Exportar PI^S (JSON)"); self.btn_exp.setObjectName("btnSuccess")
        self.btn_exp.clicked.connect(self._export_pis); self.btn_exp.setEnabled(False)
        ll.addWidget(self.btn_exp)
        ll.addStretch()
        layout.addWidget(left)

        right = QVBoxLayout()
        self.prob_bar = ProbabilityBar(); right.addWidget(self.prob_bar)
        self.tbl_sc = QTableWidget(); self.tbl_sc.setColumnCount(11)
        self.tbl_sc.setHorizontalHeaderLabels(["w","Etiqueta","P(w)","Cat","Viento(kt)","D_w",
            "Lat","Lon","Lugar Impacto","Analogos","Horizonte"])
        self.tbl_sc.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.tbl_sc.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeMode.Stretch)
        self.tbl_sc.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeMode.Stretch)
        self.tbl_sc.setAlternatingRowColors(True)
        right.addWidget(self.tbl_sc)

        g = QGroupBox("Coordenadas de Impacto y Referencia de Pronostico")
        gl = QVBoxLayout(g)
        self.tbl_coords = QTableWidget(); self.tbl_coords.setColumnCount(10)
        self.tbl_coords.setHorizontalHeaderLabels(
            ["Huracan","Ano","Lat","Lon","Region","Cat","Viento(kt)","Dist.Costa(km)","Landfall","Modelo Pronostico"])
        self.tbl_coords.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeMode.Stretch)
        self.tbl_coords.setAlternatingRowColors(True); self.tbl_coords.setMaximumHeight(200)
        gl.addWidget(self.tbl_coords)
        right.addWidget(g)

        self.lbl_met = QLabel(""); self.lbl_met.setWordWrap(True)
        self.lbl_met.setStyleSheet("background:#F0F4F8;padding:8px;border-radius:4px;font-family:monospace;font-size:11px;")
        right.addWidget(self.lbl_met)
        layout.addLayout(right, 1)
        return tab

    # --- TAB 2 ---
    def _tab_cascada(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        ctrl = QHBoxLayout()
        ctrl.addWidget(QLabel("Huracan:"))
        self.combo_hc = QComboBox(); ctrl.addWidget(self.combo_hc)
        ctrl.addWidget(QLabel("Region:"))
        self.combo_rc = QComboBox(); self.combo_rc.addItems(["(Auto)","Occidente","Centro","Oriente"]); ctrl.addWidget(self.combo_rc)
        ctrl.addWidget(QLabel("Metodo:"))
        self.combo_mc = QComboBox(); self.combo_mc.addItems(["GMM (EM)","K-Means++"]); ctrl.addWidget(self.combo_mc)
        self.btn_cas = QPushButton("Generar Cascada (72-12h)"); self.btn_cas.clicked.connect(self._run_cascade)
        ctrl.addWidget(self.btn_cas)
        layout.addLayout(ctrl)
        sp = QSplitter(Qt.Orientation.Vertical)
        self.chart_conv = ChartLabel(); self.chart_conv.setMinimumHeight(300); sp.addWidget(self.chart_conv)
        self.tbl_cas = QTableWidget(); self.tbl_cas.setColumnCount(7)
        self.tbl_cas.setHorizontalHeaderLabels(["Horizonte","w","P(w)","Categoria","Viento(kt)","D_w","Entropia H"])
        self.tbl_cas.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tbl_cas.setAlternatingRowColors(True); sp.addWidget(self.tbl_cas)
        sp.setSizes([350,350]); layout.addWidget(sp, 1)
        return tab

    # --- TAB 3 ---
    def _tab_analysis(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        ctrl = QHBoxLayout()
        ctrl.addWidget(QLabel("Huracan:"))
        self.combo_ha = QComboBox(); ctrl.addWidget(self.combo_ha)
        ctrl.addWidget(QLabel("Region:"))
        self.combo_ra = QComboBox(); self.combo_ra.addItems(["(Auto)","Occidente","Centro","Oriente"]); ctrl.addWidget(self.combo_ra)
        self.btn_full = QPushButton("Ejecutar Analisis Completo"); self.btn_full.setObjectName("btnDanger")
        self.btn_full.clicked.connect(self._run_full); ctrl.addWidget(self.btn_full)
        self.btn_xlsx = QPushButton("Exportar Excel (.xlsx)"); self.btn_xlsx.setObjectName("btnSuccess")
        self.btn_xlsx.clicked.connect(self._export_excel); self.btn_xlsx.setEnabled(False); ctrl.addWidget(self.btn_xlsx)
        self.btn_jfull = QPushButton("Exportar JSON"); self.btn_jfull.setObjectName("btnWarning")
        self.btn_jfull.clicked.connect(self._export_json); self.btn_jfull.setEnabled(False); ctrl.addWidget(self.btn_jfull)
        layout.addLayout(ctrl)

        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        sc = QWidget(); grid = QGridLayout(sc); grid.setSpacing(8)
        self.chart_slots = {}
        names = [("prob","Distribucion de Probabilidad"),("bayes","Convergencia Bayesiana"),
                 ("ent","Entropia de Shannon"),("silk","Silhouette vs k"),
                 ("bick","BIC vs k"),("cmp","Comparacion de Metodos"),
                 ("heat","Mapa de Calor Categorias"),("dem","Factor de Demanda"),
                 ("impmap","Mapa de Impacto por Escenario"),("impcas","Evolucion Punto Impacto")]
        for i, (key, title) in enumerate(names):
            gb = QGroupBox(title); gbl = QVBoxLayout(gb)
            cl = ChartLabel(); cl.setMinimumSize(420, 300); gbl.addWidget(cl)
            self.chart_slots[key] = cl; grid.addWidget(gb, i//2, i%2)
        scroll.setWidget(sc); layout.addWidget(scroll, 1)
        return tab

    # --- TAB 4 ---
    def _tab_realtime(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        ctrl = QHBoxLayout()
        self.btn_nhc = QPushButton("Consultar NHC Ahora"); self.btn_nhc.setObjectName("btnDanger")
        self.btn_nhc.clicked.connect(self._run_rt); ctrl.addWidget(self.btn_nhc)
        self.lbl_nhc = QLabel("No consultado"); ctrl.addWidget(self.lbl_nhc); ctrl.addStretch()
        layout.addLayout(ctrl)
        self.tbl_nhc = QTableWidget(); self.tbl_nhc.setColumnCount(8)
        self.tbl_nhc.setHorizontalHeaderLabels(["Nombre","Tipo","Cat","Lat","Lon","Viento(kt)","Dist Cuba(km)","Amenaza"])
        self.tbl_nhc.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.tbl_nhc, 1)
        self.txt_adv = QTextEdit(); self.txt_adv.setReadOnly(True); self.txt_adv.setMaximumHeight(150)
        self.txt_adv.setPlaceholderText("Avisos NHC apareceran aqui.\nTemporada: Jun 1 - Nov 30.")
        layout.addWidget(self.txt_adv)
        return tab

    # --- TAB 5 ---
    def _tab_catalogo(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        sl = QHBoxLayout()
        cat = self.engine.storm_catalog
        lf = sum(1 for s in cat if s["landfall"])
        ymin = min(s["year"] for s in cat); ymax = max(s["year"] for s in cat)
        for label, value in [("Huracanes",str(self.engine.database_size)),("Landfalls",str(lf)),
                             ("Rango","%d-%d"%(ymin,ymax)),("Regiones","Occ|Cen|Ori")]:
            b = QGroupBox(label); bl = QVBoxLayout(b)
            l = QLabel(value); l.setObjectName("statLabel"); l.setAlignment(Qt.AlignmentFlag.AlignCenter)
            bl.addWidget(l); sl.addWidget(b)
        layout.addLayout(sl)
        self.tbl_cat = QTableWidget(); self.tbl_cat.setColumnCount(8)
        self.tbl_cat.setHorizontalHeaderLabels(["Ano","Nombre","Cat","Region","Landfall","Dist(km)","Viento(kt)","Provincias"])
        self.tbl_cat.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.Stretch)
        self.tbl_cat.setAlternatingRowColors(True); self._fill_cat()
        layout.addWidget(self.tbl_cat, 1)
        return tab

    # --- TAB 6: BASE DE DATOS ---
    def _tab_database(self):
        tab = QWidget(); layout = QVBoxLayout(tab)

        # --- Section 1: Status ---
        grp_status = QGroupBox("Estado de la Base de Datos")
        sl = QVBoxLayout(grp_status)
        self.lbl_db_status = QLabel("")
        self.lbl_db_status.setWordWrap(True)
        self.lbl_db_status.setStyleSheet(
            "background:#F8F9FA;border:1px solid #BDC3C7;border-radius:4px;padding:10px;font-size:11px;")
        self._refresh_db_status()
        sl.addWidget(self.lbl_db_status)
        layout.addWidget(grp_status)

        # --- Section 2: Sync NHC ---
        grp_sync = QGroupBox("Sincronizar con NHC (HURDAT2 Oficial)")
        sync_l = QVBoxLayout(grp_sync)
        sync_l.addWidget(QLabel(
            "Descarga el archivo HURDAT2 oficial del National Hurricane Center, "
            "filtra huracanes que afectan Cuba, y fusiona con la base actual. "
            "Los duplicados se ignoran automaticamente."))
        sync_btn_l = QHBoxLayout()
        self.btn_sync = QPushButton("Sincronizar NHC (Internet)")
        self.btn_sync.setObjectName("btnDanger"); self.btn_sync.clicked.connect(self._do_sync)
        sync_btn_l.addWidget(self.btn_sync)
        self.btn_save_db = QPushButton("Guardar BD Local")
        self.btn_save_db.setObjectName("btnSuccess"); self.btn_save_db.clicked.connect(self._do_save_db)
        sync_btn_l.addWidget(self.btn_save_db)
        self.btn_load_db = QPushButton("Cargar BD Local")
        self.btn_load_db.clicked.connect(self._do_load_db)
        sync_btn_l.addWidget(self.btn_load_db)
        sync_btn_l.addStretch()
        sync_l.addLayout(sync_btn_l)
        self.txt_sync_log = QTextEdit(); self.txt_sync_log.setReadOnly(True)
        self.txt_sync_log.setMaximumHeight(120)
        self.txt_sync_log.setPlaceholderText("Los resultados de sincronizacion apareceran aqui...")
        sync_l.addWidget(self.txt_sync_log)
        layout.addWidget(grp_sync)

        # --- Section 3: Manual entry ---
        grp_manual = QGroupBox("Agregar Huracan Manualmente")
        ml = QVBoxLayout(grp_manual)
        ml.addWidget(QLabel(
            "Para huracanes recientes aun no publicados en HURDAT2 (temporada actual). "
            "Ingrese los datos del track en formato: YYYYMMDD, HHMM, STATUS, LAT, LON, WIND_KT, PRESS_MB"))
        form = QGridLayout()
        form.addWidget(QLabel("ATCF ID:"), 0, 0)
        self.txt_atcf = QComboBox(); self.txt_atcf.setEditable(True)
        self.txt_atcf.setCurrentText("AL__2025")
        form.addWidget(self.txt_atcf, 0, 1)
        form.addWidget(QLabel("Nombre:"), 0, 2)
        self.txt_storm_name = QComboBox(); self.txt_storm_name.setEditable(True)
        form.addWidget(self.txt_storm_name, 0, 3)
        ml.addLayout(form)
        ml.addWidget(QLabel("Track points (una linea por punto, separados por coma):"))
        self.txt_track_points = QTextEdit()
        self.txt_track_points.setMaximumHeight(160)
        self.txt_track_points.setPlaceholderText(
            "20250912, 0000, HU, 18.0, -74.0, 85, 978\n"
            "20250912, 1200, HU, 19.0, -75.5, 95, 970\n"
            "20250913, 0000, HU, 20.0, -77.0, 110, 958\n"
            "20250913, 1200, HU, 20.8, -78.0, 120, 950\n"
            "20250914, 0000, HU, 21.5, -79.0, 125, 945\n"
            "20250914, 1200, HU, 22.5, -80.0, 115, 952\n"
            "20250915, 0000, HU, 23.5, -81.0, 100, 962\n"
            "20250915, 1200, TS, 24.5, -82.0, 60, 985")
        ml.addWidget(self.txt_track_points)
        btn_add_l = QHBoxLayout()
        self.btn_add_storm = QPushButton("Agregar a la Base de Datos")
        self.btn_add_storm.setObjectName("btnSuccess"); self.btn_add_storm.clicked.connect(self._do_add_storm)
        btn_add_l.addWidget(self.btn_add_storm)
        btn_add_l.addStretch()
        ml.addLayout(btn_add_l)
        self.txt_add_log = QTextEdit(); self.txt_add_log.setReadOnly(True)
        self.txt_add_log.setMaximumHeight(100)
        self.txt_add_log.setPlaceholderText("Resultado de la operacion...")
        ml.addWidget(self.txt_add_log)
        layout.addWidget(grp_manual)

        layout.addStretch()
        return tab

    def _refresh_db_status(self):
        rpt = self.engine.db_report()
        yr = rpt.get("year_range", (0, 0))
        regions = rpt.get("regions", {})
        cats = rpt.get("categories", {})
        cat_str = ", ".join("Cat%d: %d" % (c, n) for c, n in sorted(cats.items()))
        reg_str = ", ".join("%s: %d" % (r, n) for r, n in regions.items())
        db_path = DatabaseManager.DEFAULT_DB_PATH
        local_exists = os.path.exists(db_path)
        local_info = ""
        if local_exists:
            import json as _j
            with open(db_path, "r") as _f:
                meta = _j.load(_f).get("metadata", {})
            local_info = "  Archivo local: %s (actualizado: %s)" % (
                db_path, meta.get("last_sync", meta.get("saved", "?")))
        else:
            local_info = "  Sin archivo local (usando base embebida)"
        self.lbl_db_status.setText(
            "<b>Huracanes sobre Cuba:</b> %d<br>"
            "<b>Rango:</b> %d - %d<br>"
            "<b>Regiones:</b> %s<br>"
            "<b>Categorias:</b> %s<br>"
            "<b>Landfalls:</b> %d<br>"
            "<b>Persistencia:</b> %s" % (
                rpt["n_storms"], yr[0], yr[1], reg_str, cat_str,
                rpt["landfalls"], local_info))

    def _do_sync(self):
        self.btn_sync.setEnabled(False)
        self.txt_sync_log.setPlainText("Descargando HURDAT2 de NHC...\nURL: %s\n" %
            DatabaseManager.get_nhc_hurdat2_url())
        QApplication.processEvents()
        try:
            rpt = self.engine.sync_nhc()
            if rpt.get("error"):
                self.txt_sync_log.append("ERROR: %s" % rpt["error"])
            else:
                self.txt_sync_log.append("Sincronizacion exitosa:")
                self.txt_sync_log.append("  Tormentas en HURDAT2: %d" % rpt["nhc_total"])
                self.txt_sync_log.append("  Afectan Cuba: %d" % rpt["nhc_cuba"])
                self.txt_sync_log.append("  Nuevas agregadas: %d" % rpt["new_added"])
                self.txt_sync_log.append("  Duplicadas ignoradas: %d" % rpt["duplicates_skipped"])
                self.txt_sync_log.append("  BD total ahora: %d" % rpt.get("db_size_after", "?"))
                self.txt_sync_log.append("  Guardado en: %s" % rpt.get("saved_to", "?"))
                self._refresh_all()
        except Exception as e:
            self.txt_sync_log.append("EXCEPCION: %s" % str(e))
        self.btn_sync.setEnabled(True)

    def _do_save_db(self):
        fp, _ = QFileDialog.getSaveFileName(self, "Guardar BD",
            DatabaseManager.DEFAULT_DB_PATH, "JSON (*.json)")
        if not fp: return
        try:
            msg = self.engine.save_local_db(fp)
            self.txt_sync_log.setPlainText(msg)
            self._refresh_db_status()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _do_load_db(self):
        fp, _ = QFileDialog.getOpenFileName(self, "Cargar BD", "", "JSON (*.json)")
        if not fp: return
        try:
            msg = self.engine.load_local_db(fp)
            self.txt_sync_log.setPlainText(msg)
            self._refresh_all()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _do_add_storm(self):
        atcf = self.txt_atcf.currentText().strip()
        name = self.txt_storm_name.currentText().strip()
        if not atcf or not name:
            self.txt_add_log.setPlainText("ERROR: ATCF ID y Nombre son obligatorios")
            return
        raw = self.txt_track_points.toPlainText().strip()
        if not raw:
            self.txt_add_log.setPlainText("ERROR: Ingrese al menos 2 puntos de track")
            return
        track_points = []
        for line in raw.split("\n"):
            line = line.strip()
            if not line: continue
            parts = [p.strip() for p in line.split(",")]
            if len(parts) != 7:
                self.txt_add_log.setPlainText(
                    "ERROR en linea: %s\nSe esperan 7 campos separados por coma" % line)
                return
            try:
                track_points.append((
                    parts[0], parts[1], parts[2],
                    float(parts[3]), float(parts[4]),
                    int(parts[5]), int(parts[6]),
                ))
            except ValueError as e:
                self.txt_add_log.setPlainText("ERROR parseando: %s\n%s" % (line, e))
                return
        try:
            msg = self.engine.add_storm(atcf, name, track_points)
            self.txt_add_log.setPlainText(msg)
            if msg.startswith("OK") or msg.startswith("ADVERTENCIA"):
                self._refresh_all()
        except Exception as e:
            self.txt_add_log.setPlainText("EXCEPCION: %s" % str(e))

    def _refresh_all(self):
        """Refresh all combos and tables after DB change."""
        for c in [self.combo_huracan, self.combo_hc, self.combo_ha]: c.clear()
        self._populate_storms()
        self._fill_cat()
        self._refresh_db_status()
        self.statusBar().showMessage("BD actualizada: %d huracanes" % self.engine.database_size)

    # --- POPULATE ---
    def _populate_storms(self):
        items = []
        for s in self.engine.storm_catalog:
            items.append("%s (%d) - Cat %d %s" % (s['name'], s['year'], s['category'], s['region']))
        for c in [self.combo_huracan, self.combo_hc, self.combo_ha]:
            c.addItems(items)
        if items: self._on_hchange(0)

    def _on_hchange(self, idx):
        if idx < 0: return
        cat = self.engine.storm_catalog
        if idx >= len(cat): return
        s = cat[idx]; imp = self.engine.generator.impacts[idx]
        lf_txt = '<span style="color:red;">LANDFALL</span>' if s['landfall'] else 'Sin landfall'
        prov = ', '.join(s['provinces'][:4])
        prov_etc = '...' if len(s['provinces']) > 4 else ''
        self.lbl_info.setText(
            "<b>%s (%d)</b><br>Cat: <b>%d</b> (%d kt)<br>Region: <b>%s</b><br>"
            "%s - %.0f km<br>Lat: <b>%.2f</b> Lon: <b>%.2f</b><br>"
            "Prov: %s%s<br>Transito: %.0fh | Rumbo: %.0f deg<br>"
            "<b>HURDAT2 Best Track (NHC)</b>" % (
                s['name'], s['year'], s['category'], s['wind_kt'], s['region'],
                lf_txt, s['closest_km'], imp.closest_point.lat, imp.closest_point.lon,
                prov, prov_etc, imp.transit_hours, imp.entry_bearing))

    def _fill_cat(self):
        cat = self.engine.storm_catalog; self.tbl_cat.setRowCount(len(cat))
        for i, s in enumerate(cat):
            self.tbl_cat.setItem(i, 0, QTableWidgetItem(str(s["year"])))
            self.tbl_cat.setItem(i, 1, QTableWidgetItem(s["name"]))
            ci = QTableWidgetItem(category_label(s["category"]))
            ci.setBackground(QColor(CAT_COLORS.get(s["category"], "#CCC")))
            self.tbl_cat.setItem(i, 2, ci)
            self.tbl_cat.setItem(i, 3, QTableWidgetItem(s["region"]))
            self.tbl_cat.setItem(i, 4, QTableWidgetItem("SI" if s["landfall"] else ""))
            self.tbl_cat.setItem(i, 5, QTableWidgetItem("%.0f" % s['closest_km']))
            self.tbl_cat.setItem(i, 6, QTableWidgetItem(str(s["wind_kt"])))
            self.tbl_cat.setItem(i, 7, QTableWidgetItem(", ".join(s["provinces"][:3])))

    # --- HELPERS ---
    def _sname(self, c):
        t = c.currentText(); return t.split(" (")[0] if " (" in t else t
    def _reg(self, c):
        r = c.currentText(); return None if r == "(Auto)" else r
    def _met(self, c):
        return "gmm" if "GMM" in c.currentText() else "kmeans"
    def _hor(self):
        return int(self.combo_h.currentText().replace("h",""))

    # --- ACTIONS ---
    def _run_analysis(self):
        p = {"hurricane_name": self._sname(self.combo_huracan), "horizon": self._hor(),
             "region": self._reg(self.combo_region), "method": self._met(self.combo_met)}
        n = self.spin_n.value()
        if n > 0: p["n_scenarios"] = n
        self._start("analyze", p)

    def _run_cascade(self):
        self._start("cascade", {"hurricane_name": self._sname(self.combo_hc),
                                 "region": self._reg(self.combo_rc), "method": self._met(self.combo_mc)})

    def _run_full(self):
        self._start("full_analysis", {"hurricane_name": self._sname(self.combo_ha),
                                       "region": self._reg(self.combo_ra)})

    def _run_rt(self):
        self._start("realtime", {})

    def _start(self, task, params):
        self.progress.setVisible(True); self.progress.setValue(0)
        for b in [self.btn_gen, self.btn_cas, self.btn_full]: b.setEnabled(False)
        self.worker = WorkerThread(self.engine, task, params)
        self.worker.progress.connect(lambda m,v: (self.progress.setValue(v), self.statusBar().showMessage(m)))
        self.worker.finished.connect(self._done)
        self.worker.error.connect(self._err)
        self.worker.start()

    def _err(self, msg):
        self.progress.setVisible(False); self._enbtn()
        QMessageBox.warning(self, "Error", msg)

    def _enbtn(self):
        for b in [self.btn_gen, self.btn_cas, self.btn_full]: b.setEnabled(True)

    def _done(self, result):
        self.progress.setVisible(False); self._enbtn()
        if isinstance(result, ScenarioSet):
            self.resultado_actual = result; self._show_sc(result); self._show_coords(result)
            self.btn_exp.setEnabled(True)
        elif isinstance(result, dict) and result.get("type") == "full_analysis":
            self.full_analysis_data = result; self.cascada_actual = result["cascade"]
            self._show_cascade(result["cascade"]); self._show_charts(result)
            self.btn_xlsx.setEnabled(True); self.btn_jfull.setEnabled(True)
        elif isinstance(result, dict) and all(isinstance(v, ScenarioSet) for v in result.values()):
            self.cascada_actual = result; self._show_cascade(result)
        elif isinstance(result, list):
            self._show_rt(result)

    # --- DISPLAY ---
    def _show_sc(self, ss):
        self.prob_bar.set_data(ss.scenarios)
        self.tbl_sc.setRowCount(len(ss.scenarios))
        for i, s in enumerate(ss.scenarios):
            self.tbl_sc.setItem(i, 0, QTableWidgetItem("w%d" % s.id))
            self.tbl_sc.setItem(i, 1, QTableWidgetItem(s.label))
            pi = QTableWidgetItem("%.4f" % s.probability)
            pi.setTextAlignment(Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter)
            self.tbl_sc.setItem(i, 2, pi)
            ci = QTableWidgetItem(category_label(s.category))
            ci.setBackground(QColor(CAT_COLORS.get(s.category, "#CCC")))
            self.tbl_sc.setItem(i, 3, ci)
            self.tbl_sc.setItem(i, 4, QTableWidgetItem(str(s.max_wind_kt)))
            di = QTableWidgetItem("%.3f" % s.demand_factor)
            di.setTextAlignment(Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter)
            self.tbl_sc.setItem(i, 5, di)
            # Impact coordinates and location
            self.tbl_sc.setItem(i, 6, QTableWidgetItem("%.4f" % s.impact_lat))
            self.tbl_sc.setItem(i, 7, QTableWidgetItem("%.4f" % s.impact_lon))
            loc_txt = s.impact_location
            if s.impact_province:
                loc_txt += " (%s)" % s.impact_province
            self.tbl_sc.setItem(i, 8, QTableWidgetItem(loc_txt))
            self.tbl_sc.setItem(i, 9, QTableWidgetItem(", ".join(s.reference_storms[:3])))
            self.tbl_sc.setItem(i, 10, QTableWidgetItem("%dh" % s.horizon_hours))
        ent = -sum(s.probability*math.log2(s.probability) for s in ss.scenarios if s.probability > 0)
        bic_txt = "BIC: %.1f | " % ss.bic if ss.bic else ""
        self.lbl_met.setText(
            "%s (%d) | %s | %dh | %s | Analogos: %d | Sil: %.3f | %sH=%.3f bits | SumP=%.6f" % (
                ss.hurricane_name, ss.hurricane_year, ss.target_region.value,
                ss.horizon_hours, ss.clustering_method, ss.n_analogs_used,
                ss.silhouette_score, bic_txt, ent, ss.probability_sum))

    def _show_coords(self, ss):
        target = None
        for imp in self.engine.generator.impacts:
            if imp.track.name.upper() == ss.hurricane_name.upper():
                target = imp; break
        analogs = self.engine.generator.find_analogs(ss.hurricane_name, n_analogs=15)
        all_imp = []
        if target: all_imp.append((">>> " + target.track.name, target, True))
        for a, sim in analogs: all_imp.append((a.track.name, a, False))
        self.tbl_coords.setRowCount(len(all_imp))
        for i, (name, imp, is_t) in enumerate(all_imp):
            self.tbl_coords.setItem(i, 0, QTableWidgetItem(name))
            self.tbl_coords.setItem(i, 1, QTableWidgetItem(str(imp.track.year)))
            self.tbl_coords.setItem(i, 2, QTableWidgetItem("%.4f" % imp.closest_point.lat))
            self.tbl_coords.setItem(i, 3, QTableWidgetItem("%.4f" % imp.closest_point.lon))
            self.tbl_coords.setItem(i, 4, QTableWidgetItem(imp.region.value))
            ci = QTableWidgetItem(category_label(imp.category_at_closest))
            ci.setBackground(QColor(CAT_COLORS.get(imp.category_at_closest, "#CCC")))
            self.tbl_coords.setItem(i, 5, ci)
            self.tbl_coords.setItem(i, 6, QTableWidgetItem(str(imp.wind_at_closest)))
            self.tbl_coords.setItem(i, 7, QTableWidgetItem("%.1f" % imp.closest_approach_km))
            self.tbl_coords.setItem(i, 8, QTableWidgetItem("SI" if imp.landfall else "NO"))
            self.tbl_coords.setItem(i, 9, QTableWidgetItem("HURDAT2 Best Track (NHC)"))
            if is_t:
                for c in range(10):
                    cell = self.tbl_coords.item(i, c)
                    if cell:
                        cell.setBackground(QColor("#D5E8D4"))
                        cell.setFont(QFont("Arial", 9, QFont.Weight.Bold))

    def _show_cascade(self, cascade):
        fig = Charts.bayesian(cascade); self.chart_conv.set_figure(fig)
        rows = []
        for h in sorted(cascade.keys(), reverse=True):
            ss = cascade[h]
            ent = -sum(s.probability*math.log2(s.probability) for s in ss.scenarios if s.probability > 1e-15)
            for s in ss.scenarios: rows.append((h, s, ent))
        self.tbl_cas.setRowCount(len(rows))
        for i, (h, s, ent) in enumerate(rows):
            hi = QTableWidgetItem("%dh" % h); hi.setBackground(QColor(HORIZON_COLORS.get(h, "#CCC")))
            self.tbl_cas.setItem(i, 0, hi)
            self.tbl_cas.setItem(i, 1, QTableWidgetItem("w%d" % s.id))
            pi = QTableWidgetItem("%.4f" % s.probability)
            pi.setTextAlignment(Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter)
            self.tbl_cas.setItem(i, 2, pi)
            ci = QTableWidgetItem(category_label(s.category))
            ci.setBackground(QColor(CAT_COLORS.get(s.category, "#CCC")))
            self.tbl_cas.setItem(i, 3, ci)
            self.tbl_cas.setItem(i, 4, QTableWidgetItem(str(s.max_wind_kt)))
            self.tbl_cas.setItem(i, 5, QTableWidgetItem("%.3f" % s.demand_factor))
            self.tbl_cas.setItem(i, 6, QTableWidgetItem("%.3f" % ent))

    def _show_charts(self, data):
        cascade = data["cascade"]; compare = data["compare"]
        k_sil = data["k_silhouettes"]; k_bic = data["k_bics"]
        ss48 = cascade.get(48, list(cascade.values())[0])
        coast = None
        if ENGINE_OK:
            coast = CUBA_COASTLINE_POINTS
        figs = {"prob": Charts.prob_dist(ss48), "bayes": Charts.bayesian(cascade),
                "ent": Charts.entropy(cascade), "silk": Charts.sil_k(k_sil),
                "bick": Charts.bic_k(k_bic), "cmp": Charts.method_cmp(compare),
                "heat": Charts.cat_heat(cascade), "dem": Charts.demand(cascade),
                "impmap": Charts.impact_map(ss48, coast),
                "impcas": Charts.impact_cascade(cascade, coast)}
        for key, fig in figs.items():
            if key in self.chart_slots: self.chart_slots[key].set_figure(fig)
        self.tabs.setCurrentIndex(2)
        self.statusBar().showMessage("Analisis completo: 10 graficos. Listo para exportar.")

    def _show_rt(self, storms):
        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        if not storms:
            self.lbl_nhc.setText("Sin amenazas - %s" % now)
            self.lbl_nhc.setStyleSheet("color:#27AE60;font-weight:bold;")
            self.tbl_nhc.setRowCount(0)
            self.txt_adv.setPlainText("NHC: %s\nSin ciclones activos amenazando Cuba." % now)
            return
        self.lbl_nhc.setText("%d amenaza(s) - %s" % (len(storms), now))
        self.lbl_nhc.setStyleSheet("color:#E74C3C;font-weight:bold;")
        self.tbl_nhc.setRowCount(len(storms))
        for i, s in enumerate(storms):
            self.tbl_nhc.setItem(i, 0, QTableWidgetItem(s.get("name","?")))
            self.tbl_nhc.setItem(i, 1, QTableWidgetItem(s.get("type","?")))
            cat = s.get("category", -1)
            ci = QTableWidgetItem(category_label(cat))
            ci.setBackground(QColor(CAT_COLORS.get(cat, "#CCC")))
            self.tbl_nhc.setItem(i, 2, ci)
            self.tbl_nhc.setItem(i, 3, QTableWidgetItem("%.1f" % s.get("lat",0)))
            self.tbl_nhc.setItem(i, 4, QTableWidgetItem("%.1f" % s.get("lon",0)))
            self.tbl_nhc.setItem(i, 5, QTableWidgetItem(str(s.get("wind_kt",0))))
            self.tbl_nhc.setItem(i, 6, QTableWidgetItem(str(s.get("distance_to_cuba_km","?"))))
            thr = s.get("threat_level","?")
            ti = QTableWidgetItem(thr)
            tc = {"ALTO":"#E74C3C","MEDIO":"#F39C12","BAJO":"#27AE60"}.get(thr, "#CCC")
            ti.setBackground(QColor(tc)); ti.setForeground(QColor("white"))
            self.tbl_nhc.setItem(i, 7, ti)

    # --- EXPORTS ---
    def _export_pis(self):
        if not self.resultado_actual:
            QMessageBox.information(self, "Info", "Primero genere escenarios."); return
        fp, _ = QFileDialog.getSaveFileName(self, "Exportar PI^S", "scenarios_PIS.json", "JSON (*.json)")
        if not fp: return
        ss = self.resultado_actual
        d = {"n_scenarios": ss.n_scenarios, "hurricane": ss.hurricane_name,
             "year": ss.hurricane_year, "horizon_hours": ss.horizon_hours,
             "region": ss.target_region.value,
             "scenarios": [s.to_dict() for s in ss.scenarios]}
        with open(fp, "w", encoding="utf-8") as f: json.dump(d, f, indent=2, ensure_ascii=False)
        self.statusBar().showMessage("Exportado: %s" % fp)

    def _export_excel(self):
        if not self.full_analysis_data:
            QMessageBox.information(self, "Info", "Primero ejecute Analisis Completo (pestana 3)."); return
        if not EXCEL_OK:
            QMessageBox.warning(self, "Error", "Instale openpyxl: pip install openpyxl"); return
        fp, _ = QFileDialog.getSaveFileName(self, "Exportar Excel", "hurricane_analysis.xlsx", "Excel (*.xlsx)")
        if not fp: return
        try:
            ExcelExporter.export(fp, self.full_analysis_data)
            self.statusBar().showMessage("Excel: %s (9 hojas)" % fp)
            QMessageBox.information(self, "OK",
                "Excel generado con 9 hojas:\n1.Resumen\n2.Escenarios\n3.Convergencia\n"
                "4.Comparacion\n5.Seleccion k\n6.Analogos\n7.Coordenadas\n8.Metricas ML\n"
                "9.Impacto por Escenario\n\n%s" % fp)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _export_json(self):
        if not self.full_analysis_data:
            QMessageBox.information(self, "Info", "Ejecute Analisis Completo."); return
        fp, _ = QFileDialog.getSaveFileName(self, "JSON", "full_analysis.json", "JSON (*.json)")
        if not fp: return
        d = self.full_analysis_data
        out = {"cascade": {str(h): ss.to_dict() for h, ss in d["cascade"].items()},
               "compare": {m: ss.to_dict() for m, ss in d["compare"].items()},
               "k_silhouettes": d["k_silhouettes"],
               "k_bics": {str(k): v for k, v in d["k_bics"].items()},
               "analog_details": d["analog_details"]}
        with open(fp, "w", encoding="utf-8") as f: json.dump(out, f, indent=2, ensure_ascii=False, default=str)
        self.statusBar().showMessage("JSON: %s" % fp)

    def _load_hurdat2(self):
        fp, _ = QFileDialog.getOpenFileName(self, "HURDAT2", "", "Text (*.txt);;All (*)")
        if not fp: return
        try:
            self.engine = HurricaneScenarioEngine(hurdat2_path=fp)
            for c in [self.combo_huracan, self.combo_hc, self.combo_ha]: c.clear()
            self._populate_storms(); self._fill_cat()
            self.statusBar().showMessage("HURDAT2: %d huracanes" % self.engine.database_size)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))


def main():
    app = QApplication(sys.argv); app.setStyle("Fusion")
    p = app.palette()
    p.setColor(QPalette.ColorRole.Window, QColor("#F0F4F8"))
    p.setColor(QPalette.ColorRole.WindowText, QColor("#2C3E50"))
    app.setPalette(p)
    w = VentanaPrincipal(); w.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
