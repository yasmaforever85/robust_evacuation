#!/usr/bin/env python3
"""
═══════════════════════════════════════════════════════════════════════════════
VALIDATION V6 — MIP-based stochastic superiority + Figure generation
═══════════════════════════════════════════════════════════════════════════════

Extends V5 with:
  - Formal EV/EEV/VSS terminology (Birge & Louveaux, 2011)
  - Uncertainty characterization figure (4-panel)
  - All computational figures (V1–V5) regenerated from real data
  - Bayesian cascade extraction for theoretical verification

100 experiments: 5 network instances × 4 hurricanes × 5 forecast horizons.

Instances:
  G1: Havana (base)        — 2A, 1R, 2F — P_nom=60,   Σπ=75
  G2: Havana extended      — 4A, 2R, 3F — P_nom=150,  Σπ=180
  G3: Western Cuba         — 6A, 3R, 5F — P_nom=350,  Σπ=413
  G4: Multi-province       — 10A,5R, 8F — P_nom=800,  Σπ=920
  G5: National-scale       — 15A,6R,10F — P_nom=1500, Σπ=1680

Dependencies:
  - GeneradorEscenariosHuracan.py (scenario engine with Eq. 9 demand factor)
  - PI_Estoc_Esc.py (two-stage stochastic MIP)
  - PI_Plan_Flujo.py (deterministic flow MIP)
  - ortools (OR-Tools SCIP solver)
  - matplotlib, numpy, openpyxl

Output:
  - Console summary
  - VMSTA_validation_results.xlsx (3 sheets: All_Experiments, Instance_Summary, Metadata)
  - img/ directory with all figures for JUS paper

Terminology (§2.5.1 of JUS paper):
  - P^EV : Expected-value problem — deterministic model with mean demand
  - EEV  : Expected result of EV solution — P^EV plan evaluated under scenarios
  - VSS  : Value of the Stochastic Solution = EEV − Z*_{P^S}
  - P^S  : Two-stage stochastic program with recourse

@author: Yasmany Fernández Fernández (Asere)
@date: 2025-02-15
@version: 6.0
═══════════════════════════════════════════════════════════════════════════════
"""

import sys, os, math, json, random, warnings, io, time
import numpy as np
warnings.filterwarnings("ignore")

# ── Path setup ──
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from GeneradorEscenariosHuracan import ScenarioGenerator
from PI_Estoc_Esc import ModeloEstocasticoPI
from PI_Plan_Flujo import ModeloAIMMS, generar_idf
from ortools.linear_solver import pywraplp

np.random.seed(42)
random.seed(42)

# ═══════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════
THETA_PLUS  = 10000.0   # penalty: unmet demand (per person)
THETA_MINUS = 100.0     # penalty: excess evacuation (per person)
C_TRANSPORT = 1.0       # transport cost per person-km
KAPPA       = 1.3       # road adjustment factor
EARTH_R     = 6371.0    # km

HURRICANES = ["IAN", "IRMA", "SANDY", "MATTHEW"]
HORIZONS   = [72, 48, 36, 24, 12]

INSTANCES = [
    {'name': 'G1', 'P_nom': 60,   'cap_ratio': 1.25,
     'nA': 2,  'nR': 1, 'nF': 2, 'region': 'Occidente'},
    {'name': 'G2', 'P_nom': 150,  'cap_ratio': 1.20,
     'nA': 4,  'nR': 2, 'nF': 3, 'region': 'Occidente'},
    {'name': 'G3', 'P_nom': 350,  'cap_ratio': 1.18,
     'nA': 6,  'nR': 3, 'nF': 5, 'region': 'Occidente'},
    {'name': 'G4', 'P_nom': 800,  'cap_ratio': 1.15,
     'nA': 10, 'nR': 5, 'nF': 8, 'region': 'Centro'},
    {'name': 'G5', 'P_nom': 1500, 'cap_ratio': 1.12,
     'nA': 15, 'nR': 6, 'nF': 10, 'region': 'Centro'},
]

CUBA_REGIONS = {
    'Occidente': {'lat': (22.0, 23.2), 'lon': (-84.0, -81.5)},
    'Centro':    {'lat': (21.5, 22.8), 'lon': (-81.5, -79.0)},
}


# ═══════════════════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════════════════
def haversine(lat1, lon1, lat2, lon2):
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp, dl = math.radians(lat2 - lat1), math.radians(lon2 - lon1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * EARTH_R * math.asin(math.sqrt(a))


# ═══════════════════════════════════════════════════════════════════════
#  NETWORK BUILDER
# ═══════════════════════════════════════════════════════════════════════
def build_network(inst):
    """Build a network instance compatible with PI_Estoc_Esc."""
    nA, nR, nF = inst['nA'], inst['nR'], inst['nF']
    region = inst['region']
    P_nom = inst['P_nom']
    cap_ratio = inst['cap_ratio']

    if inst['name'] == 'G1':
        from PI_Estoc_Esc import crear_red_simple_compatible_pii as _g1_net
        _ns, _nt, _nl, _arcos, _coords, _caps, _fi = _g1_net()
        return {
            'name': 'G1',
            'nodos_salida': _ns, 'nodos_transito': _nt, 'nodos_llegada': _nl,
            'arcos': _arcos, 'coordenadas': _coords, 'capacidades': _caps,
            'fi_nominal': _fi,
            'P_nom': sum(h*q for (h,_),q in _fi.items()),
            'Sigma_pi': sum(v for k,v in _caps.items() if k[0]=='pi'),
            'nodes': len(_ns)+len(_nt)+len(_nl), 'n_arcs': len(_arcos),
        }
    else:
        lr = CUBA_REGIONS[region]
        origins  = [(np.random.uniform(*lr['lat']), np.random.uniform(*lr['lon']))
                     for _ in range(nA)]
        transits = [(np.random.uniform(*lr['lat']), np.random.uniform(*lr['lon']))
                     for _ in range(nR)]
        shelters = [(np.random.uniform(*lr['lat']), np.random.uniform(*lr['lon']))
                     for _ in range(nF)]

    nodos_salida  = [f'A{i+1}' for i in range(nA)]
    nodos_transito = [f'R{j+1}' for j in range(nR)]
    nodos_llegada = [f'F{k+1}' for k in range(nF)]

    coordenadas = {}
    for i, c in enumerate(origins):  coordenadas[f'A{i+1}'] = c
    for j, c in enumerate(transits): coordenadas[f'R{j+1}'] = c
    for k, c in enumerate(shelters): coordenadas[f'F{k+1}'] = c

    arcos = {}
    for i, o in enumerate(origins):
        dists = [(j, haversine(o[0], o[1], t[0], t[1]))
                 for j, t in enumerate(transits)]
        dists.sort(key=lambda x: x[1])
        for j, d in dists[:min(2, nR)]:
            arcos[(f'A{i+1}', f'R{j+1}')] = round(d * KAPPA, 2)
    for j, t in enumerate(transits):
        for k, s in enumerate(shelters):
            d = haversine(t[0], t[1], s[0], s[1])
            arcos[(f'R{j+1}', f'F{k+1}')] = round(d * KAPPA, 2)

    Sigma_pi = int(round(P_nom * cap_ratio))
    pi_base = Sigma_pi // nF
    pi_list = [pi_base] * nF
    pi_list[-1] += Sigma_pi - sum(pi_list)

    capacidades = {}
    for j in range(nR):
        capacidades[('beta', f'R{j+1}')] = 2 * Sigma_pi
    for k in range(nF):
        capacidades[('pi', f'F{k+1}')] = pi_list[k]
        capacidades[('gamma', f'F{k+1}')] = 2 * Sigma_pi

    fam_sizes = [3, 4, 5]
    fi_nominal = {}
    persons_per_origin = P_nom // nA
    remainder = P_nom - persons_per_origin * nA
    for i in range(nA):
        target = persons_per_origin + (1 if i < remainder else 0)
        allocated = 0
        for h_idx, h in enumerate(fam_sizes):
            if h_idx < len(fam_sizes) - 1:
                n_fam = max(1, target // (len(fam_sizes) * h))
                allocated += n_fam * h
            else:
                n_fam = max(1, (target - allocated) // h)
            fi_nominal[(h, f'A{i+1}')] = n_fam

    P_actual = sum(h * q for (h, _), q in fi_nominal.items())
    return {
        'name': inst['name'],
        'nodos_salida': nodos_salida, 'nodos_transito': nodos_transito,
        'nodos_llegada': nodos_llegada, 'arcos': arcos,
        'coordenadas': coordenadas, 'capacidades': capacidades,
        'fi_nominal': fi_nominal, 'P_nom': P_actual,
        'Sigma_pi': Sigma_pi, 'nodes': nA + nR + nF, 'n_arcs': len(arcos),
    }


# ═══════════════════════════════════════════════════════════════════════
#  SOLVER WRAPPERS
# ═══════════════════════════════════════════════════════════════════════
class _ModeloEstocasticoFixed(ModeloEstocasticoPI):
    """Patched: uses element-wise MAX fi across scenarios for IDF."""
    def _generar_idf_unificado(self):
        from PI_Plan_Flujo import generar_idf as _gen_idf
        fi_max = {}
        for esc in self.scenarios:
            for k, v in esc['fi'].items():
                fi_max[k] = max(fi_max.get(k, 0), v)
        idf_max_dict = _gen_idf(fi_max)
        idf_set = set(idf_max_dict.keys())
        cantidades = {}
        for s, esc in enumerate(self.scenarios):
            cantidades[s] = {}
            for (id_fam, h, ns) in idf_set:
                cantidades[s][(id_fam, h, ns)] = esc['fi'].get((h, ns), 0)
        return idf_set, cantidades, idf_max_dict


def solve_stochastic_MIP(net, scenarios_raw):
    """Solve P^S using real two-stage MIP."""
    model_scenarios = []
    for idx, sc in enumerate(scenarios_raw):
        fi_s = {k: max(0, int(round(v * sc['eta'])))
                for k, v in net['fi_nominal'].items()}
        demanda_personas = {}
        for (h, nodo), cantidad in fi_s.items():
            demanda_personas[nodo] = demanda_personas.get(nodo, 0) + h * cantidad
        model_scenarios.append({
            'id': idx, 'prob': sc['prob'], 'fi': fi_s,
            'demanda_personas': demanda_personas,
            'desc': f'ω{idx+1}', 'tipo': 'gmm_bayesian'
        })

    old = sys.stdout; sys.stdout = io.StringIO()
    try:
        modelo = _ModeloEstocasticoFixed()
        modelo.set_coordenadas(net['coordenadas'])
        modelo.scenarios = model_scenarios
        modelo.num_scenarios = len(model_scenarios)
        t0 = time.time()
        sol = modelo.resolver_estocastico(
            net['nodos_salida'], net['nodos_transito'], net['nodos_llegada'],
            net['arcos'], net['capacidades'],
            c=C_TRANSPORT, theta_plus=THETA_PLUS, theta_minus=THETA_MINUS,
            cobertura_minima=0.0, cobertura_minima_obligatoria=None,
            verbose=False
        )
        solve_time = time.time() - t0
    except Exception:
        sys.stdout = old
        return None
    finally:
        sys.stdout = old

    if not sol:
        return None
    Z_ps = sol['costo_total']
    dp = sol.get('delta_plus', {})
    max_deficit = 0
    min_coverage = 100.0
    for s_idx, sc in enumerate(model_scenarios):
        def_s = sum(v for (si, _), v in dp.items() if si == s_idx)
        dem_s = sum(sc['demanda_personas'].values())
        cov_s = (1 - def_s / max(dem_s, 1)) * 100
        if def_s > max_deficit: max_deficit = int(def_s)
        if cov_s < min_coverage: min_coverage = cov_s
    return Z_ps, max_deficit, round(min_coverage, 1), solve_time


def solve_deterministic_EV(net, scenarios_raw):
    """
    Solve P^EV (expected-value problem) and compute EEV.

    P^EV: deterministic model with expected demand d̄ = Σ p_s · d_s
    EEV:  cost of P^EV solution evaluated under all scenarios
    VSS:  EEV − Z*_{P^S}  (Birge & Louveaux, 2011)
    """
    fi_nom = net['fi_nominal']
    Sigma_pi = net['Sigma_pi']

    # Expected fi
    fi_ev = {}
    for k in fi_nom.keys():
        fi_ev[k] = max(0, int(round(
            sum(sc['prob'] * sc['eta'] * fi_nom[k] for sc in scenarios_raw)
        )))
    P_ev = sum(h * q for (h, _), q in fi_ev.items())

    if P_ev > Sigma_pi:
        return None, False, None, P_ev, 0.0

    old = sys.stdout; sys.stdout = io.StringIO()
    try:
        idf_ev = generar_idf(fi_ev)
        modelo_ev = ModeloAIMMS(
            net['nodos_salida'], net['nodos_transito'], net['nodos_llegada'],
            net['arcos'], idf_ev, net['capacidades']
        )
        modelo_ev.crear_variables()
        modelo_ev.agregar_restricciones()
        costo = sum(
            C_TRANSPORT * net['arcos'].get((i, j), 0) * h *
            modelo_ev.X[(id_val, h, i, j)]
            for (id_val, h, i, j) in modelo_ev.X.keys()
            if (i, j) in net['arcos']
        )
        modelo_ev.solver.Minimize(costo)
        status = modelo_ev.solver.Solve()
    except Exception:
        sys.stdout = old
        return None, False, None, P_ev, 0.0
    finally:
        sys.stdout = old

    if status not in [pywraplp.Solver.OPTIMAL, pywraplp.Solver.FEASIBLE]:
        return None, False, None, P_ev, 0.0

    Z_ev = modelo_ev.solver.Objective().Value()

    # EEV: evaluate EV plan under each scenario
    total_eev = 0.0
    max_deficit_ev = 0
    min_coverage_ev = 100.0
    for sc in scenarios_raw:
        P_s = int(round(sum(h * max(0, int(round(fi_nom[(h, ns)] * sc['eta'])))
                            for (h, ns) in fi_nom.keys())))
        deficit = max(0, P_s - P_ev)
        deficit = max(deficit, max(0, P_s - Sigma_pi))
        excess = max(0, P_ev - P_s)
        recourse = THETA_PLUS * deficit + THETA_MINUS * excess
        eev_s = Z_ev + recourse
        total_eev += sc['prob'] * eev_s
        if deficit > max_deficit_ev: max_deficit_ev = deficit
        cov = (1 - deficit / max(P_s, 1)) * 100
        if cov < min_coverage_ev: min_coverage_ev = cov

    return Z_ev, True, total_eev, max_deficit_ev, round(min_coverage_ev, 1)


# ═══════════════════════════════════════════════════════════════════════
#  MAIN EXPERIMENT LOOP
# ═══════════════════════════════════════════════════════════════════════
def run_experiments():
    gen = ScenarioGenerator()
    networks = {inst['name']: build_network(inst) for inst in INSTANCES}
    all_results = []
    summaries = []

    # Also collect scenario cascade data for figures
    cascade_data = {}   # {hurricane: {horizon: {scenarios, entropy, ...}}}
    all_scenarios_72 = {}  # {hurricane: scenario data at 72h}

    for inst_name, net in networks.items():
        print(f"\n{'='*60}")
        print(f"  {inst_name}: P_nom={net['P_nom']}, Σπ={net['Sigma_pi']}, "
              f"{net['nodes']} nodes, {net['n_arcs']} arcs")
        print(f"{'='*60}")

        n_infeasible = 0
        vss_values = []
        cov_ps_values = []
        def_ps_values = []

        for hurr in HURRICANES:
            for h in HORIZONS:
                try:
                    ss = gen.generate_scenarios(hurr, h, method="gmm", seed=42)
                except Exception:
                    continue

                scenarios_raw = [
                    {'prob': s.probability, 'eta': s.demand_factor,
                     'cat': s.category, 'wind': s.max_wind_kt,
                     'location': s.impact_location}
                    for s in ss.scenarios
                ]

                # Collect cascade data for G1 figures
                if inst_name == 'G1':
                    if hurr not in cascade_data:
                        cascade_data[hurr] = {}
                    probs = [sc['prob'] for sc in scenarios_raw]
                    etas = [sc['eta'] for sc in scenarios_raw]
                    entropy = -sum(p * np.log2(p) for p in probs if p > 0)
                    n_active = sum(1 for p in probs if p > 0.01)
                    P_s_list = [int(round(e * net['P_nom'])) for e in etas]
                    cascade_data[hurr][h] = {
                        'scenarios': scenarios_raw,
                        'probs': probs, 'etas': etas,
                        'demands': P_s_list,
                        'entropy': entropy, 'n_active': n_active,
                    }
                    if h == 72:
                        P_ev_val = int(round(sum(
                            p * e * net['P_nom']
                            for p, e in zip(probs, etas))))
                        all_scenarios_72[hurr] = {
                            'probs': probs, 'demands': P_s_list,
                            'entropy': entropy, 'P_ev': P_ev_val
                        }

                # Solve P^S
                ps_result = solve_stochastic_MIP(net, scenarios_raw)
                if ps_result is None:
                    print(f"  {hurr} {h}h: SOLVER FAILED")
                    continue
                Z_ps, def_ps, cov_ps, t_solve = ps_result

                # Solve P^EV and compute EEV
                P_ev_data = solve_deterministic_EV(net, scenarios_raw)
                Z_ev, det_feas, eev, def_ev, cov_ev = P_ev_data

                P_ev = int(round(sum(
                    sc['prob'] * sc['eta'] * net['P_nom']
                    for sc in scenarios_raw)))

                if det_feas and eev and eev > 0:
                    vss = eev - Z_ps
                    vss_pct = vss / eev * 100
                    if vss_pct < 0: vss_pct = 0.0
                else:
                    vss = float('inf')
                    vss_pct = float('inf')
                    n_infeasible += 1

                status = '✓' if det_feas else '✗'
                vss_str = f"{vss_pct:.1f}%" if vss_pct != float('inf') else "INF"
                print(f"  {hurr:8s} {h:2d}h: {status} P_ev={P_ev:4d} "
                      f"VSS={vss_str:>7s} Cov={cov_ps:.1f}% "
                      f"Def={def_ps:3d} [{t_solve:.1f}s]")

                row = {
                    'Instance': inst_name, 'Hurricane': hurr,
                    'Horizon (h)': h,
                    'P_ev': P_ev, 'Σπ': net['Sigma_pi'],
                    'Det. Feasible': '✓' if det_feas else '✗',
                    'Z_PS': round(Z_ps, 2),
                    'Z_EV': round(Z_ev, 2) if Z_ev else '—',
                    'EEV': round(eev, 2) if eev else '—',
                    'VSS (%)': round(vss_pct, 1) if vss_pct != float('inf') else '—',
                    'Deficit PS': def_ps,
                    'Deficit EV': def_ev if def_ev else '—',
                    'Coverage PS (%)': cov_ps,
                    'Coverage EV (%)': cov_ev if cov_ev else '—',
                }
                all_results.append(row)

                if vss != float('inf') and vss_pct > 0:
                    vss_values.append(vss_pct)
                cov_ps_values.append(cov_ps)
                def_ps_values.append(def_ps)

        summaries.append({
            'Instance': inst_name,
            'Nodes': net['nodes'], 'Arcs': net['n_arcs'],
            'P_nom': net['P_nom'], 'Σπ': net['Sigma_pi'],
            'Experiments': 20, 'Infeasible': n_infeasible,
            'Infeas. (%)': round(n_infeasible / 20 * 100, 0),
            'Mean Cov. PS (%)': round(np.mean(cov_ps_values), 1) if cov_ps_values else 0,
            'Min Cov. PS (%)': round(min(cov_ps_values), 1) if cov_ps_values else 0,
            'Mean VSS (%)': round(np.mean(vss_values), 1) if vss_values else '—',
            'Max Deficit': max(def_ps_values) if def_ps_values else 0,
        })

    return all_results, summaries, cascade_data, all_scenarios_72


# ═══════════════════════════════════════════════════════════════════════
#  EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════
def write_xlsx(results, summaries, path):
    """Write results to Excel with proper column names."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True)
    hdr_fill = PatternFill('solid', fgColor='D9E1F2')
    thin = Border(Side('thin'), Side('thin'), Side('thin'), Side('thin'))

    # Sheet 1: All experiments
    ws1 = wb.active
    ws1.title = 'All_Experiments'
    headers = ['Instance', 'Hurricane', 'Horizon (h)', 'P_ev', 'Σπ',
               'Det. Feasible', 'Z_PS', 'Z_EV', 'EEV', 'VSS (%)',
               'Deficit PS', 'Deficit EV', 'Coverage PS (%)', 'Coverage EV (%)']
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = thin
    for i, r in enumerate(results, 2):
        for c, h in enumerate(headers, 1):
            cell = ws1.cell(row=i, column=c, value=r[h])
            cell.border = thin

    # Sheet 2: Summary
    ws2 = wb.create_sheet('Instance_Summary')
    s_headers = list(summaries[0].keys())
    for c, h in enumerate(s_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = thin
    for i, s in enumerate(summaries, 2):
        for c, h in enumerate(s_headers, 1):
            cell = ws2.cell(row=i, column=c, value=s[h])
            cell.border = thin

    # Sheet 3: Metadata
    ws3 = wb.create_sheet('Metadata')
    meta = [
        ('JUS Validation Results — Real MIP Solver (OR-Tools SCIP)', None),
        (None, None),
        ('Model', 'Two-stage stochastic program with recourse (P^S)'),
        ('Solver', 'OR-Tools SCIP via PI_Estoc_Esc.py'),
        ('Demand factor', 'Eq. 9 calibrated: α₀=0.36, α_cat=0.22, α_wind=0.08, α_dist=0.10, α_land=0.12'),
        ('Scenario generation', 'GMM-Bayesian via GeneradorEscenariosHuracan.py'),
        ('θ⁺ (deficit penalty)', f'{THETA_PLUS:,.0f} $/person'),
        ('θ⁻ (excess penalty)', f'{THETA_MINUS:,.0f} $/person'),
        ('Transport cost', f'{C_TRANSPORT} $/person-km'),
        (None, None),
        ('Terminology', None),
        ('P^EV', 'Expected-value problem: deterministic model with mean demand'),
        ('EEV', 'Expected result of EV solution: P^EV plan under scenarios'),
        ('VSS', 'Value of the Stochastic Solution = EEV − Z*_{P^S}'),
        ('P^S', 'Two-stage stochastic program with recourse'),
        ('Reference', 'Birge & Louveaux (2011), Introduction to Stochastic Programming'),
    ]
    for i, (k, v) in enumerate(meta, 1):
        if k: ws3.cell(row=i, column=1, value=k)
        if v: ws3.cell(row=i, column=2, value=v)

    wb.save(path)
    print(f"\n📊 Results saved to {path}")


# ═══════════════════════════════════════════════════════════════════════
#  FIGURE GENERATION
# ═══════════════════════════════════════════════════════════════════════
def generate_all_figures(results, cascade_data, all_scenarios_72, img_dir='img'):
    """Generate all figures for the JUS paper from real data."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    os.makedirs(img_dir, exist_ok=True)

    HCOLORS = {'IAN': '#E74C3C', 'IRMA': '#3498DB',
               'SANDY': '#2ECC71', 'MATTHEW': '#9B59B6'}
    HMARKERS = {'IAN': 's', 'IRMA': 'o', 'SANDY': '^', 'MATTHEW': 'D'}
    INST_COLORS = ['#E74C3C', '#3498DB', '#2ECC71', '#F39C12', '#9B59B6']
    instances = ['G1', 'G2', 'G3', 'G4', 'G5']
    cap_ratios = [1.25, 1.20, 1.18, 1.15, 1.12]

    # ─── FIG UNCERTAINTY ANALYSIS (4-panel) ───
    fig, axes = plt.subplots(2, 2, figsize=(14, 11))
    ian = cascade_data.get('IAN', {})
    Sigma_pi = 75

    # (a) Bayesian probability cascade
    ax = axes[0, 0]
    sc_labels = ['$\\omega_1$ (Cat 3)', '$\\omega_2$ (Cat 4)',
                 '$\\omega_3$ (Cat 5)', '$\\omega_4$ (Cat 2)', '$\\omega_5$ (Cat 4)']
    sc_colors = ['#E74C3C', '#3498DB', '#2ECC71', '#F39C12', '#9B59B6']
    for i in range(min(5, len(ian.get(72, {}).get('probs', [])))):
        probs = [ian[h]['probs'][i] for h in HORIZONS if h in ian]
        hs = [h for h in HORIZONS if h in ian]
        lbl = sc_labels[i] if i < len(sc_labels) else f'$\\omega_{i+1}$'
        ax.plot(hs, probs, 'o-', color=sc_colors[i % 5], label=lbl,
                markersize=8, linewidth=2.5)
    ax.set_xlabel('Forecast horizon $\\tau$ (hours)', fontsize=12)
    ax.set_ylabel('Posterior probability $p_s^{(\\tau)}$', fontsize=12)
    ax.set_title('(a) Bayesian probability cascade — Hurricane Ian', fontsize=13)
    ax.legend(fontsize=9, ncol=2, loc='center right')
    ax.invert_xaxis(); ax.set_ylim(-0.02, 1.02); ax.grid(True, alpha=0.3)

    # (b) Uncertainty set vs capacity
    ax = axes[0, 1]
    for hurr, hdata in all_scenarios_72.items():
        pairs = sorted(zip(hdata['demands'], hdata['probs']))
        ax.scatter([p[0] for p in pairs], [p[1] for p in pairs],
                   s=120, color=HCOLORS[hurr], marker='o',
                   edgecolors='black', linewidth=0.5, zorder=5,
                   label=f'{hurr} ($P_{{ev}}$={hdata["P_ev"]})')
        ax.plot([p[0] for p in pairs], [p[1] for p in pairs],
                '--', color=HCOLORS[hurr], alpha=0.3)
    ax.axvline(x=Sigma_pi, color='red', linewidth=2.5, alpha=0.7,
               label=f'$\\Sigma\\pi = {Sigma_pi}$')
    ax.fill_betweenx([0, 0.7], Sigma_pi, 100, alpha=0.08, color='red')
    ax.text(Sigma_pi + 1, 0.62, 'Infeasible\nregion', fontsize=10, color='red', alpha=0.7)
    ax.set_xlabel('Scenario demand $P_s$ (persons)', fontsize=12)
    ax.set_ylabel('Scenario probability $p_s$', fontsize=12)
    ax.set_title('(b) Uncertainty set $\\Xi^{(72)}$ vs. capacity $\\Sigma\\pi$', fontsize=13)
    ax.legend(fontsize=8, loc='upper left')
    ax.set_xlim(40, 95); ax.set_ylim(-0.02, 0.7); ax.grid(True, alpha=0.3)

    # (c) Entropy + VSS
    ax = axes[1, 0]
    entropies = [ian[h]['entropy'] for h in HORIZONS if h in ian]
    hs_plot = [h for h in HORIZONS if h in ian]
    ax.plot(hs_plot, entropies, 's-', color='#E74C3C', markersize=10,
            linewidth=2.5, label='Shannon entropy $H(\\tau)$')
    ax2 = ax.twinx()
    vss_vals = []
    for r in results:
        if r['Instance'] == 'G1' and r['Hurricane'] == 'IAN':
            v = r['VSS (%)']
            vss_vals.append(float(v) if v != '—' else 0)
    if len(vss_vals) == len(hs_plot):
        ax2.bar(hs_plot, vss_vals, width=4, alpha=0.3, color='#3498DB',
                edgecolor='#3498DB', linewidth=1, label='VSS (\\%)')
    ax2.set_ylabel('VSS / EEV (\\%)', fontsize=12, color='#3498DB')
    ax2.tick_params(axis='y', labelcolor='#3498DB'); ax2.set_ylim(0, 60)
    ax.set_xlabel('Forecast horizon $\\tau$ (hours)', fontsize=12)
    ax.set_ylabel('Shannon entropy $H$ (bits)', fontsize=12, color='#E74C3C')
    ax.tick_params(axis='y', labelcolor='#E74C3C')
    ax.set_title('(c) Entropy reduction and VSS trajectory', fontsize=13)
    ax.invert_xaxis(); ax.set_ylim(0, 2.0); ax.grid(True, alpha=0.3)
    lines1, labels1 = ax.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax.legend(lines1 + lines2, labels1 + labels2, fontsize=10, loc='upper left')
    if len(entropies) > 1:
        pct = (1 - entropies[-1] / entropies[0]) * 100
        ax.annotate(f'$\\Delta H = {pct:.0f}\\%$ reduction',
                    xy=(hs_plot[-1], entropies[-1]), xytext=(30, 1.2),
                    arrowprops=dict(arrowstyle='->', color='#E74C3C'),
                    fontsize=11, color='#E74C3C', fontweight='bold')

    # (d) Recourse landscape
    ax = axes[1, 1]
    if 72 in ian:
        r_vals = np.arange(40, 95, 1)
        P_nom = 60
        for i, sc in enumerate(ian[72]['scenarios']):
            P_s = int(round(sc['eta'] * P_nom))
            costs = []
            for r in r_vals:
                deficit = max(0, P_s - r)
                excess = max(0, r - P_s)
                costs.append(sc['prob'] * (THETA_PLUS * deficit + THETA_MINUS * excess) / 1000)
            ax.plot(r_vals, costs, '-', color=sc_colors[i % 5], linewidth=2,
                    label=f'$p_{i+1} Q(r, \\omega_{i+1})$, $P_{i+1}$={P_s}', alpha=0.8)
        E_Q = []
        for r in r_vals:
            eq = sum(sc['prob'] * (THETA_PLUS * max(0, int(round(sc['eta']*P_nom)) - r) +
                     THETA_MINUS * max(0, r - int(round(sc['eta']*P_nom))))
                     for sc in ian[72]['scenarios'])
            E_Q.append(eq)
        ax.plot(r_vals, [c/1000 for c in E_Q], 'k-', linewidth=3,
                label='$\\mathcal{Q}(r) = \\sum_s p_s Q(r, \\omega_s)$')
        r_opt = r_vals[np.argmin(E_Q)]
        ax.axvline(x=r_opt, color='green', linestyle='--', alpha=0.7, linewidth=1.5)
        ax.text(r_opt + 1, max(c/1000 for c in E_Q) * 0.85,
                f'$r^* = {r_opt}$', color='green', fontsize=11)
        ax.axvline(x=Sigma_pi, color='red', linestyle='-', alpha=0.5, linewidth=1.5)
        ax.text(Sigma_pi + 1, max(c/1000 for c in E_Q) * 0.95,
                f'$\\Sigma\\pi$', color='red', fontsize=11)
    ax.set_xlabel('First-stage allocation $r$ (persons)', fontsize=12)
    ax.set_ylabel('Prob.-weighted recourse (\\$×10³)', fontsize=12)
    ax.set_title('(d) Expected recourse landscape $\\mathcal{Q}(r)$', fontsize=13)
    ax.legend(fontsize=8, loc='upper right'); ax.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(os.path.join(img_dir, 'fig_uncertainty_analysis.pdf'),
                bbox_inches='tight', dpi=150)
    plt.close()
    print(f"  ✓ fig_uncertainty_analysis.pdf")

    # ─── FIG V1: VSS across horizons ───
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
    for hurr in HURRICANES:
        rows = [r for r in results if r['Instance'] == 'G1' and r['Hurricane'] == hurr]
        rows.sort(key=lambda x: -x['Horizon (h)'])
        hs = [r['Horizon (h)'] for r in rows]
        vss_abs = [float(r['EEV']) - float(r['Z_PS'])
                   if r['EEV'] != '—' and r['VSS (%)'] != '—'
                   and float(r['VSS (%)']) > 0 else None for r in rows]
        vss_pct = [float(r['VSS (%)']) if r['VSS (%)'] != '—'
                   and float(r['VSS (%)']) > 0 else None for r in rows]
        vh = [h for h, v in zip(hs, vss_abs) if v is not None]
        vv = [v for v in vss_abs if v is not None]
        if vv:
            ax1.plot(vh, vv, '-' + HMARKERS[hurr], color=HCOLORS[hurr],
                     label=hurr, markersize=8, linewidth=2)
        vh2 = [h for h, v in zip(hs, vss_pct) if v is not None]
        vv2 = [v for v in vss_pct if v is not None]
        if vv2:
            ax2.plot(vh2, vv2, '-' + HMARKERS[hurr], color=HCOLORS[hurr],
                     label=hurr, markersize=8, linewidth=2)
        for r in rows:
            if r['Det. Feasible'] == '✗':
                ax2.axvspan(r['Horizon (h)'] - 3, r['Horizon (h)'] + 3,
                            alpha=0.08, color=HCOLORS[hurr])
    ax1.set_xlabel('Forecast horizon $h$ (hours)'); ax1.set_ylabel('VSS (\\$)')
    ax1.set_title('(a) Value of the Stochastic Solution')
    ax1.legend(); ax1.invert_xaxis(); ax1.grid(True, alpha=0.3)
    ax2.set_xlabel('Forecast horizon $h$ (hours)'); ax2.set_ylabel('VSS / EEV (\\%)')
    ax2.set_title('(b) Relative VSS (shaded = $P^{EV}$ infeasible)')
    ax2.legend(); ax2.invert_xaxis(); ax2.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(img_dir, 'figV1_vss_horizons.pdf'), bbox_inches='tight', dpi=150)
    plt.close()
    print(f"  ✓ figV1_vss_horizons.pdf")

    # ─── FIG V2: Coverage & Deficit ───
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
    for hurr in HURRICANES:
        rows = [r for r in results if r['Instance'] == 'G1' and r['Hurricane'] == hurr]
        rows.sort(key=lambda x: -x['Horizon (h)'])
        hs = [r['Horizon (h)'] for r in rows]
        cov_ps = [float(r['Coverage PS (%)']) for r in rows]
        def_ps = [int(r['Deficit PS']) for r in rows]
        ax1.plot(hs, cov_ps, '-' + HMARKERS[hurr], color=HCOLORS[hurr],
                 label=f'{hurr} ($P^S$)', markersize=8, linewidth=2)
        ax2.plot(hs, def_ps, '-' + HMARKERS[hurr], color=HCOLORS[hurr],
                 label=f'{hurr} ($P^S$)', markersize=8, linewidth=2)
        # EV dashed
        for r in rows:
            if r['Det. Feasible'] == '✓' and r['Coverage EV (%)'] != '—':
                ax1.plot(r['Horizon (h)'], float(r['Coverage EV (%)']),
                         HMARKERS[hurr], color=HCOLORS[hurr],
                         markersize=6, alpha=0.4, markerfacecolor='none')
    ax1.set_xlabel('Forecast horizon $h$ (hours)'); ax1.set_ylabel('Worst-case coverage (\\%)')
    ax1.set_title('(a) $P^S$ (solid) vs EV (open markers)')
    ax1.legend(fontsize=8, ncol=2); ax1.invert_xaxis(); ax1.grid(True, alpha=0.3)
    ax2.set_xlabel('Forecast horizon $h$ (hours)'); ax2.set_ylabel('Max unmet demand (persons)')
    ax2.set_title('(b) Deficit: $P^S$')
    ax2.legend(fontsize=8, ncol=2); ax2.invert_xaxis(); ax2.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(img_dir, 'figV2_coverage_deficit.pdf'), bbox_inches='tight', dpi=150)
    plt.close()
    print(f"  ✓ figV2_coverage_deficit.pdf")

    # ─── FIG V3: Feasibility coverage bars ───
    fig, ax = plt.subplots(figsize=(12, 5))
    x_pos = np.arange(len(HORIZONS))
    width = 0.15
    for idx, inst in enumerate(instances):
        covs = []
        hatches_list = []
        for h in HORIZONS:
            rows = [r for r in results if r['Instance'] == inst and r['Horizon (h)'] == h]
            covs.append(np.mean([float(r['Coverage PS (%)']) for r in rows]) if rows else 0)
            n_inf = sum(1 for r in rows if r['Det. Feasible'] == '✗')
            hatches_list.append(n_inf >= 3)
        bars = ax.bar(x_pos + idx * width, covs, width, label=inst,
                      color=INST_COLORS[idx], alpha=0.8, edgecolor='black', linewidth=0.5)
        for bar, hatch in zip(bars, hatches_list):
            if hatch: bar.set_hatch('//')
    ax.set_xlabel('Forecast horizon'); ax.set_ylabel('$P^S$ worst-case coverage (\\%)')
    ax.set_title('Stochastic model coverage (hatched = $P^{EV}$ infeasible)')
    ax.set_xticks(x_pos + 2 * width); ax.set_xticklabels([f'{h}h' for h in HORIZONS])
    ax.legend(); ax.set_ylim(60, 100); ax.grid(True, alpha=0.3, axis='y')
    plt.tight_layout()
    plt.savefig(os.path.join(img_dir, 'figV3_feasibility_coverage.pdf'), bbox_inches='tight', dpi=150)
    plt.close()
    print(f"  ✓ figV3_feasibility_coverage.pdf")

    # ─── FIG V5: Scalability (3 panels) ───
    fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(16, 5))
    infeas_rates, mean_covs, min_covs, mean_vss_list = [], [], [], []
    for inst in instances:
        rows = [r for r in results if r['Instance'] == inst]
        n_inf = sum(1 for r in rows if r['Det. Feasible'] == '✗')
        infeas_rates.append(n_inf / len(rows) * 100 if rows else 0)
        mean_covs.append(np.mean([float(r['Coverage PS (%)']) for r in rows]) if rows else 0)
        min_covs.append(min(float(r['Coverage PS (%)']) for r in rows) if rows else 0)
        vss_v = [float(r['VSS (%)']) for r in rows
                 if r['VSS (%)'] != '—' and float(r['VSS (%)']) > 0]
        mean_vss_list.append(np.mean(vss_v) if vss_v else 0)

    bars = ax1.bar(instances, infeas_rates, color=INST_COLORS, edgecolor='black', linewidth=0.5)
    for bar, rate in zip(bars, infeas_rates):
        ax1.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1,
                 f'{rate:.0f}%', ha='center', fontsize=11, fontweight='bold')
    ax1.set_xlabel('Instance'); ax1.set_ylabel('$P^{EV}$ infeasibility rate (\\%)')
    ax1.set_title('(a) Deterministic infeasibility'); ax1.set_ylim(0, 110)
    ax1.grid(True, alpha=0.3, axis='y')

    ax2.plot(instances, mean_covs, 's-', color='#2ECC71', markersize=10, linewidth=2,
             label='Mean Cov($P^S$)')
    ax2.plot(instances, min_covs, 'v--', color='#E74C3C', markersize=10, linewidth=2,
             label='Min Cov($P^S$)')
    ax2.set_xlabel('Instance'); ax2.set_ylabel('Coverage (\\%)')
    ax2.set_title('(b) Stochastic coverage ($P^S$)'); ax2.legend()
    ax2.set_ylim(65, 100); ax2.grid(True, alpha=0.3)

    ax3.bar(instances, mean_vss_list, color=[c + '99' for c in INST_COLORS],
            edgecolor='black', linewidth=0.5, label='Mean VSS (%)')
    ax3.set_xlabel('Instance'); ax3.set_ylabel('Mean VSS (\\%)')
    ax3.set_title('(c) VSS and capacity ratio')
    ax3b = ax3.twinx()
    ax3b.plot(instances, cap_ratios, 'ko--', markersize=8, linewidth=2,
              label='$\\Sigma\\pi/P_{\\mathrm{nom}}$')
    ax3b.set_ylabel('Capacity ratio'); ax3b.set_ylim(1.08, 1.28)
    lines1, labels1 = ax3.get_legend_handles_labels()
    lines2, labels2 = ax3b.get_legend_handles_labels()
    ax3.legend(lines1 + lines2, labels1 + labels2, fontsize=9, loc='upper right')
    ax3.grid(True, alpha=0.3, axis='y')

    plt.tight_layout()
    plt.savefig(os.path.join(img_dir, 'figV5_scalability.pdf'), bbox_inches='tight', dpi=150)
    plt.close()
    print(f"  ✓ figV5_scalability.pdf")

    print(f"\n📈 All figures saved to {img_dir}/")


# ═══════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    print("=" * 60)
    print("  JUS VALIDATION V6 — Real MIP Solver (OR-Tools SCIP)")
    print("  Two-stage stochastic programming with recourse")
    print("  EV/EEV/VSS terminology (Birge & Louveaux, 2011)")
    print("=" * 60)

    t_start = time.time()
    results, summaries, cascade_data, all_scenarios_72 = run_experiments()
    t_total = time.time() - t_start

    # Print summary
    print(f"\n{'='*60}")
    print(f"  SUMMARY")
    print(f"{'='*60}")
    print(f"{'Inst':4s} {'Nodes':>5s} {'Arcs':>4s} {'Infeas%':>7s} "
          f"{'MeanCov':>7s} {'MinCov':>6s} {'MeanVSS':>7s} {'MaxDef':>6s}")
    print("-" * 52)
    for s in summaries:
        mvss = f"{s['Mean VSS (%)']:.1f}%" if isinstance(s['Mean VSS (%)'], (int, float)) else "N/A"
        print(f"{s['Instance']:4s} {s['Nodes']:5d} {s['Arcs']:4d} "
              f"{s['Infeas. (%)']:6.0f}% {s['Mean Cov. PS (%)']:6.1f}% "
              f"{s['Min Cov. PS (%)']:5.1f}% {mvss:>7s} {s['Max Deficit']:5d}")

    all_infeas = sum(1 for r in results if r['Det. Feasible'] == '✗')
    print(f"\nOverall: {all_infeas}/{len(results)} P^EV infeasible "
          f"({all_infeas / max(len(results), 1) * 100:.0f}%)")
    print(f"Total time: {t_total:.0f}s ({t_total / 60:.1f} min)")

    # Write xlsx
    xlsx_path = os.path.join(SCRIPT_DIR, 'VMSTA_validation_results.xlsx')
    write_xlsx(results, summaries, xlsx_path)

    # Generate figures
    img_dir = os.path.join(SCRIPT_DIR, 'img')
    print(f"\n{'='*60}")
    print(f"  GENERATING FIGURES")
    print(f"{'='*60}")
    generate_all_figures(results, cascade_data, all_scenarios_72, img_dir)

    print(f"\n{'='*60}")
    print(f"  DONE — All outputs ready for JUS paper")
    print(f"{'='*60}")
